"""Page 2: Variants & Analysis — per-company mapping editor.

Each row in the table = one company × one QB account.
Clicking "Save Mappings" ALWAYS shows a profile-chooser dialog
so the user explicitly picks (or creates) a profile every time.
"""

import io
import hashlib
import json
import uuid
import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from lib.master_builder import build_master_workbook, build_variants_digest, copy_sheet_into_workbook
from lib.linked_builder import build_linked_workbook
from lib.mapping_rules import map_pnl, map_bs
from lib.template_discovery import discover_template
from lib import db as dblib
from lib import profiles as P
from lib.ui import hide_streamlit_elements


# Options shared by the Template Row Formula Overrides section and the
# per-account "Formula Override" column in the main mapping table below.
PIVOT_OVERRIDE_OPTIONS = [
    "",
    "P&L Pivot CY",
    "P&L Pivot PY",
    "P&L Pivot Change",
    "BS Pivot CY",
    "BS Pivot PY",
    "BS Pivot Change",
]


# ── Cached helpers (re-run only when inputs change) ───────────────────────────
def _dict_hash(d: dict) -> str:
    return hashlib.md5(json.dumps(d, sort_keys=True, default=str).encode()).hexdigest()


@st.cache_data(show_spinner=False, max_entries=8)
def _cached_master_workbook(qb_data_hash: str, overrides_hash: str, entity_hash: str,
                             _qb_data, _overrides, _entity_overrides):
    """Build master workbook. Cache key is the triple hash; values are passed directly."""
    return build_master_workbook(
        _qb_data,
        mapping_overrides=_overrides,
        entity_mapping_overrides=_entity_overrides,
    )


@st.cache_data(show_spinner=False, max_entries=8)
def _cached_variants_digest(qb_data_hash: str, overrides_hash: str, entity_hash: str,
                             _qb_data, _pnl_pivot, _bs_pivot):
    """Build variants digest. Cache key is the triple hash."""
    return build_variants_digest(_qb_data, _pnl_pivot, _bs_pivot)


@st.cache_data(show_spinner=False, max_entries=4)
def _cached_template_lines(target_bytes_hash: str, _target_bytes: bytes):
    """Discover P&L and BS target lines from template. Cached by content hash."""
    pnl_lines: list = []
    bs_lines: list  = []
    try:
        wb_tpl = openpyxl.load_workbook(io.BytesIO(_target_bytes), data_only=False)
        year_sheets = discover_template(wb_tpl)
        for stmt in ["IS", "BS"]:
            relevant = sorted([s for s in year_sheets if s.statement == stmt],
                              key=lambda s: -(s.year or 0))
            if not relevant:
                continue
            labels = [
                r.label for r in relevant[0].rows
                if r.role in ("data", "preloaded", "subtotal")
                and r.label.strip()
            ]
            if stmt == "IS":
                pnl_lines = labels
            else:
                bs_lines = labels
    except Exception:
        pass
    return pnl_lines, bs_lines


@st.cache_data(show_spinner=False, max_entries=4)
def _cached_linked_workbook(qb_data_hash: str, target_hash: str,
                             overrides_hash: str, entity_hash: str,
                             sheets_key: str, entity_col_key: str,
                             pivot_override_key: str,
                             row_override_key: str,
                             _qb_data, _target_bytes, _overrides, _entity_overrides,
                             _selected_sheets, _entity_col_map, _sheet_pivot_overrides,
                             _row_pivot_overrides):
    """Build linked workbook. Cached by all relevant inputs."""
    return build_linked_workbook(
        _qb_data,
        _target_bytes,
        mapping_overrides=_overrides,
        entity_mapping_overrides=_entity_overrides,
        selected_sheets=_selected_sheets,
        entity_col_mapping=_entity_col_map or None,
        sheet_pivot_overrides=_sheet_pivot_overrides or None,
        row_pivot_overrides=_row_pivot_overrides or None,
    )


hide_streamlit_elements()

st.title("📊 Step 2 — Variants & Analysis")

if "qb_data" not in st.session_state:
    st.warning("Upload files first on **📂 Upload Files**.")
    st.stop()

qb_data = st.session_state.qb_data

# ── DB status ────────────────────────────────────────────────────────────────
db_online = dblib.is_connected()
if not db_online:
    st.warning(f"MongoDB unavailable — session-only mode. ({dblib.connection_error()})")

# ── Load active profile (for reading existing mappings into table) ────────────
# Auto-load most recent profile if none selected
if db_online and not st.session_state.get("active_profile_id"):
    _first = P.list_profiles()
    if _first:
        st.session_state.active_profile_id = str(_first[0]["_id"])

active_profile = None
saved_lookup: dict        = {}
entity_saved_lookup: dict = {}
if db_online:
    active_id = st.session_state.get("active_profile_id")
    if active_id:
        active_profile      = P.get_profile(active_id)
        saved_lookup        = P.mapping_lookup(active_id)
        entity_saved_lookup = P.entity_mapping_lookup(active_id)
        
        # Load row pivot overrides from database if switched profile
        if st.session_state.get("loaded_row_overrides_profile_id") != active_id:
            db_ro = {}
            for k, v in saved_lookup.items():
                if k.startswith("__template_row_override__|"):
                    parts = k.split("|", 2)
                    if len(parts) == 3:
                        db_ro[f"{parts[1]}|{parts[2]}"] = v
            st.session_state.row_pivot_overrides = db_ro
            st.session_state.loaded_row_overrides_profile_id = active_id

if active_profile:
    st.caption(f"📂 Loaded from profile: **{active_profile['name']}** — "
               f"{len(saved_lookup)} generic + {len(entity_saved_lookup)} entity-specific mappings.")

# Merge: MongoDB base + session overrides on top
session_overrides: dict        = dict(st.session_state.get("mapping_overrides", {}))
entity_session_overrides: dict = dict(st.session_state.get("entity_mapping_overrides", {}))
generic_lookup = {**saved_lookup, **session_overrides}
entity_lookup  = {**entity_saved_lookup, **entity_session_overrides}

# Compute stable cache keys for the expensive builds
_qb_hash       = _dict_hash({
    ent: {"variant": info.get("variant"), "n_pnl": len(info.get("pnl_rows", [])),
          "n_bs": len(info.get("bs_rows", []))}
    for ent, info in qb_data.items()
})
_overrides_hash = _dict_hash(generic_lookup)
_entity_hash    = _dict_hash(entity_lookup)

# ── Build master + digest (cached) ───────────────────────────────────────────
with st.spinner("Building consolidated master…"):
    master_buf, pnl_pivot, bs_pivot = _cached_master_workbook(
        _qb_hash, _overrides_hash, _entity_hash,
        qb_data, generic_lookup, entity_lookup,
    )
    digest_buf = _cached_variants_digest(
        _qb_hash, _overrides_hash, _entity_hash,
        qb_data, pnl_pivot, bs_pivot,
    )
st.session_state.pnl_pivot = pnl_pivot
st.session_state.bs_pivot  = bs_pivot


# ── Dynamic target lines from uploaded template (cached) ─────────────────────
target_lines_pnl: list = []
target_lines_bs: list  = []
template_loaded = bool(st.session_state.get("target_bytes"))
if template_loaded:
    _tb = st.session_state.target_bytes
    _tb_hash = hashlib.md5(_tb).hexdigest()
    target_lines_pnl, target_lines_bs = _cached_template_lines(_tb_hash, _tb)

# ── Build linked workbook (with pivot overrides) ──────────────────────────────
# Load saved pivot overrides from session state
_sheet_pivot_overrides: dict = st.session_state.get("sheet_pivot_overrides", {})
_row_pivot_overrides: dict = st.session_state.get("row_pivot_overrides", {})

def _is_cy_year(ys, all_sheets):
    years = [s.year for s in all_sheets if s.statement == ys.statement and s.year]
    if not years:
        return True
    return ys.year == max(years)

# ── Metrics ──────────────────────────────────────────────────────────────────
def leaf_index(pivot):
    return {k: v for k, v in pivot.items()
            if not v["is_section"] and not v["is_total"]}

pnl_leaves = leaf_index(pnl_pivot)
bs_leaves  = leaf_index(bs_pivot)
variants   = {info.get("variant", "single") for info in qb_data.values()}
variant_label = "Dual-year CY+PY" if "triple" in variants else "Single-year"

c1, c2, c3, c4 = st.columns(4)
c1.metric("Entities", len(qb_data))
c2.metric("P&L accounts", len(pnl_leaves))
c3.metric("BS accounts", len(bs_leaves))
total_abs = (sum(abs(sum(v["amounts"].values())) for v in pnl_leaves.values()) +
             sum(abs(sum(v["amounts"].values())) for v in bs_leaves.values()))
c4.metric("Total $ flowing", f"${total_abs:,.0f}")
st.caption(f"QB format: **{variant_label}**")

# ── Template sheet variables ───────────────────────────────────────────────
_target_bytes    = st.session_state.get("target_bytes")
_selected_sheets = st.session_state.get("selected_template_sheets")
_entity_col_map  = st.session_state.get("template_entity_mapping")

# Discover template sheet metadata for UI
_year_sheets_meta: list = []  # list of YearSheet objects for selected sheets
if _target_bytes and _selected_sheets:
    try:
        _wb_meta = openpyxl.load_workbook(io.BytesIO(_target_bytes), data_only=False)
        from lib.template_discovery import discover_template as _discover
        _all_ys = _discover(_wb_meta)
        _year_sheets_meta = [ys for ys in _all_ys if ys.sheet_name in _selected_sheets]
    except Exception:
        _year_sheets_meta = []

st.divider()
st.markdown("### Downloads")

# ── Pivot Tab Override UI (only when template sheets are selected) ───────────
_template_sheets_appended = False
if _target_bytes and _selected_sheets and _year_sheets_meta:
    with st.expander(
        f"⚙️ Template Row Formula Overrides ({len(_year_sheets_meta)} classified template sheet(s))",
        expanded=st.session_state.get("v2_ro_expander_expanded", False),
    ):
        st.caption(
            "By default, the formulas pull from the auto-detected Pivot tab. "
            "You can override the Pivot tab for individual template rows below. "
            "Leave blank/empty to keep the default."
        )

        selected_override_sheet = st.selectbox(
            "Select Template Sheet to view/override formulas:",
            options=_selected_sheets,
            key="v2_override_sheet_sel"
        )

        ys = next((s for s in _year_sheets_meta if s.sheet_name == selected_override_sheet), None)
        if ys:
            override_rows = []
            existing_ro = st.session_state.get("row_pivot_overrides", {})
            all_pivot_options = PIVOT_OVERRIDE_OPTIONS
            for row in ys.rows:
                if row.role not in ("data", "preloaded"):
                    continue
                sheet_base = "BS Pivot" if ys.statement == "BS" else "P&L Pivot"
                is_cy = _is_cy_year(ys, _year_sheets_meta)
                def_filter = "CY" if (ys.year is None or is_cy) else "PY"
                # Apply sheet-level override if present
                if _sheet_pivot_overrides and ys.sheet_name in _sheet_pivot_overrides:
                    def_filter = _sheet_pivot_overrides[ys.sheet_name]
                default_pivot = f"{sheet_base} {def_filter}"
                
                row_key = f"{ys.sheet_name}|{row.row_idx}"
                current_val = existing_ro.get(row_key, "")
                
                override_rows.append({
                    "row_key": row_key,
                    "Row #": row.row_idx,
                    "Template Label": row.label,
                    "Default Pivot Tab": default_pivot,
                    "Formula Pivot Tab": current_val if current_val else default_pivot,
                    "Override Pivot Tab": current_val
                })
            
            if override_rows:
                df_overrides = pd.DataFrame(override_rows)
                edited_overrides = st.data_editor(
                    df_overrides,
                    width="stretch",
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "row_key": None,
                        "Row #": st.column_config.NumberColumn(disabled=True, width="small"),
                        "Template Label": st.column_config.TextColumn(disabled=True, width="large"),
                        "Default Pivot Tab": st.column_config.TextColumn(disabled=True, width="medium"),
                        "Formula Pivot Tab": st.column_config.TextColumn(disabled=True, width="medium"),
                        "Override Pivot Tab": st.column_config.SelectboxColumn(
                            "Override Pivot Tab",
                            options=all_pivot_options,
                            width="medium",
                            required=False
                        )
                    },
                    key=f"row_overrides_editor_{selected_override_sheet}",
                    height=400
                )
                
                # Save button
                ro_save_col, _ = st.columns([2, 5])
                with ro_save_col:
                    if st.button("💾 Apply Row Overrides", type="primary", width="stretch", key="v2_row_override_save"):
                        new_ro = dict(st.session_state.get("row_pivot_overrides", {}))
                        for _, r in edited_overrides.iterrows():
                            rk = r["row_key"]
                            val = (r["Override Pivot Tab"] or "").strip()
                            if val:
                                new_ro[rk] = val
                            else:
                                new_ro.pop(rk, None)
                        st.session_state.pending_row_overrides = new_ro
                        st.session_state.v2_save_ro_dialog_open = True
                        st.session_state.v2_ro_expander_expanded = True
                        st.rerun()

                # Save Row Overrides Dialog / Confirmation
                if st.session_state.get("v2_save_ro_dialog_open"):
                    st.markdown("---")
                    with st.container(border=True):
                        st.markdown("#### 💾 Save Row Formula Overrides to Profile?")
                        st.write("Do you want to save these template row formula overrides to a profile in MongoDB so they are automatically loaded next time?")
                        
                        if db_online:
                            _profiles_now = P.list_profiles()
                            CREATE_NEW    = "__CREATE_NEW__"
                            _pids   = [str(p["_id"]) for p in _profiles_now] + [CREATE_NEW]
                            _plbls  = [p["name"]     for p in _profiles_now] + ["➕ Create new profile…"]

                            _cur = st.session_state.get("active_profile_id", "")
                            _def = _pids.index(_cur) if _cur in _pids else 0

                            sel_id = st.selectbox(
                                "Select profile to save into:",
                                options=_pids,
                                format_func=lambda x: _plbls[_pids.index(x)],
                                index=_def,
                                key="v2_ro_dialog_sel",
                            )

                            new_name = ""
                            if sel_id == CREATE_NEW:
                                new_name = st.text_input(
                                    "New profile name",
                                    placeholder="Enter a name for the new profile…",
                                    key="v2_ro_dialog_new_name",
                                )
                        else:
                            st.warning("MongoDB is offline. Overrides can only be applied to the current session.")
                            
                        c1, c2, c3 = st.columns([1.5, 1.5, 1])
                        with c1:
                            if db_online:
                                if st.button("💾 Save & Apply", type="primary", width="stretch", key="v2_ro_save_confirm"):
                                    tid = ""
                                    tname = ""
                                    if sel_id == CREATE_NEW:
                                        if not new_name.strip():
                                            st.warning("⚠️ Enter a name for the new profile first.")
                                            st.stop()
                                        tid = str(P.create_profile(new_name.strip()))
                                        tname = new_name.strip()
                                    else:
                                        tid = sel_id
                                        tname = _plbls[_pids.index(sel_id)]
                                        
                                    d = dblib.get_db()
                                    if d is not None:
                                        d.mappings.delete_many({
                                            "profile_id": P.ObjectId(tid),
                                            "statement": "__template_row_override__"
                                        })
                                        for rk_db, val_db in st.session_state.pending_row_overrides.items():
                                            if val_db:
                                                parts = rk_db.split("|", 1)
                                                if len(parts) == 2:
                                                    sheet_name, row_idx = parts[0], parts[1]
                                                    P.upsert_mapping(
                                                        profile_id=tid,
                                                        statement="__template_row_override__",
                                                        breadcrumb=sheet_name,
                                                        qb_account=row_idx,
                                                        target_line=val_db,
                                                        source="manual",
                                                    )
                                        P._invalidate_mapping_cache(tid)
                                    st.session_state.row_pivot_overrides = st.session_state.pending_row_overrides
                                    st.session_state.active_profile_id = tid
                                    st.session_state.loaded_row_overrides_profile_id = tid
                                    st.session_state.v2_save_ro_dialog_open = False
                                    st.session_state.v2_ro_expander_expanded = False
                                    st.session_state.pending_row_overrides = None
                                    st.success(f"✅ Row-level pivot overrides saved to profile **{tname}**.")
                                    st.rerun()
                            else:
                                st.button("💾 Save & Apply", type="primary", width="stretch", disabled=True, key="v2_ro_save_confirm_disabled")
                        with c2:
                            if st.button("💻 Apply to Session Only", width="stretch", key="v2_ro_apply_session"):
                                st.session_state.row_pivot_overrides = st.session_state.pending_row_overrides
                                st.session_state.v2_save_ro_dialog_open = False
                                st.session_state.v2_ro_expander_expanded = False
                                st.session_state.pending_row_overrides = None
                                st.success("✅ Row-level pivot overrides applied to session only.")
                                st.rerun()
                        with c3:
                            if st.button("❌ Cancel", width="stretch", key="v2_ro_cancel"):
                                st.session_state.v2_save_ro_dialog_open = False
                                st.session_state.v2_ro_expander_expanded = False
                                st.session_state.pending_row_overrides = None
                                st.rerun()
            else:
                st.info("No formula rows found in this sheet.")

# ── Build / append template sheets to master ────────────────────────────────
if _target_bytes and _selected_sheets:
    try:
        _t_hash  = hashlib.md5(_target_bytes).hexdigest()
        _sh_key  = json.dumps(sorted(_selected_sheets))
        _ec_key  = json.dumps(_entity_col_map, sort_keys=True, default=str) if _entity_col_map else ""
        _po_key  = json.dumps(
            {k: v for k, v in _sheet_pivot_overrides.items() if v},
            sort_keys=True,
        )
        _row_po_key = json.dumps(
            {k: v for k, v in _row_pivot_overrides.items() if v},
            sort_keys=True,
        )
        with st.spinner("Building template sheets with SUMIFS…"):
            linked_buf, _, _, _ = _cached_linked_workbook(
                _qb_hash, _t_hash, _overrides_hash, _entity_hash,
                _sh_key, _ec_key, _po_key, _row_po_key,
                qb_data, _target_bytes, generic_lookup, entity_lookup,
                _selected_sheets, _entity_col_map, _sheet_pivot_overrides,
                _row_pivot_overrides,
            )
            master_wb = openpyxl.load_workbook(io.BytesIO(master_buf.getvalue()))
            linked_wb = openpyxl.load_workbook(io.BytesIO(linked_buf.getvalue()), data_only=False)
            for sn in _selected_sheets:
                if sn.lower().strip() in ("readme", "pl long", "bs long", "coa both p&l and bs"):
                    continue
                if sn in linked_wb.sheetnames:
                    copy_sheet_into_workbook(linked_wb, sn, master_wb)
            new_buf = io.BytesIO()
            master_wb.save(new_buf)
            new_buf.seek(0)
            master_buf = new_buf
            _template_sheets_appended = True
    except Exception as _err:
        st.warning(f"⚠️ Could not append template sheets to master workbook: {_err}")

# ── Download buttons ─────────────────────────────────────────────────────────────
if _template_sheets_appended:
    _n_overrides = sum(1 for v in _row_pivot_overrides.values() if v)
    _override_note = f" • {_n_overrides} row override(s)" if _n_overrides else ""
    _master_label = (
        f"⬇ Master Consolidated Workbook "
        f"(+{len(_selected_sheets)} template sheet(s){_override_note})"
    )
else:
    _master_label = "⬇ Master Consolidated Workbook"

col1, col2 = st.columns(2)
with col1:
    st.download_button(
        _master_label,
        master_buf.getvalue(),
        "00_MASTER_Consolidated.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    if _template_sheets_appended:
        st.caption(
            f"✅ Includes: **P&L Pivot CY/PY/Change**, **BS Pivot CY/PY/Change** "
            f"+ template tabs: **{', '.join(_selected_sheets)}** with SUMIFS formulas."
        )
    elif _target_bytes and _selected_sheets is None:
        st.caption("ℹ️ Configure template sheets in **📂 Upload Files** to include them here.")
    else:
        st.caption("✅ Contains: **P&L Pivot CY/PY/Change** and **BS Pivot CY/PY/Change** pivot sheets.")
with col2:
    st.download_button(
        "⬇ Variants Digest",
        digest_buf.getvalue(),
        "01_CoA_Variants_Digest.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )


# ── Editable mapping table — per company ─────────────────────────────────────
st.divider()
st.markdown("### Map accounts to target template lines")
st.caption("Each row = one **Company × QB Account**. Set a Target Line per company. "
           "Sorted by largest $ first. Check **🧬 Duplicate?** on a row to fan that "
           "account out to a second Target Line, or **🗑 Remove?** on a duplicate row "
           "(🧬 DUPLICATE rows are marked in the Row Type column) to delete it — then "
           "click **⚡ Apply duplicate / remove** below the table.")

if template_loaded:
    st.caption(f"Dropdowns from template: {len(target_lines_pnl)} P&L lines, "
               f"{len(target_lines_bs)} BS lines.")
else:
    st.warning("⚠️ Upload a target template on **📂 Upload Files** to enable dropdowns.")


def make_rows():
    rows = []
    for entity, info in qb_data.items():
        for stmt_kind, row_key, leaves in [
            ("P&L", "pnl_rows", pnl_leaves),
            ("BS",  "bs_rows",  bs_leaves),
        ]:
            for r in info.get(row_key, []):
                if r["is_section_only"] or r["is_total"]:
                    continue
                bc, lbl = r["breadcrumb"], r["label"]

                if stmt_kind == "P&L":
                    auto, conf = map_pnl(bc, lbl)
                    if auto == "__SKIP__":
                        continue
                else:
                    auto, conf = map_bs(bc, lbl)

                entity_key  = f"E|{stmt_kind}|{entity}|{bc}|{lbl}"
                generic_key = f"{stmt_kind}|{bc}|{lbl}"

                entries     = entity_lookup.get(entity_key) or []
                primary     = next((e for e in entries if not e.get("dup_id")), None)
                dup_entries = [e for e in entries if e.get("dup_id")]

                if primary and (primary.get("target_line") or "").strip():
                    effective       = primary["target_line"]
                    pivot_override  = primary.get("pivot_override", "")
                    confidence      = "entity-saved"
                elif generic_key in generic_lookup and generic_lookup[generic_key]:
                    effective       = generic_lookup[generic_key]
                    pivot_override  = ""
                    confidence      = "saved"
                else:
                    effective       = auto or ""
                    pivot_override  = ""
                    confidence      = conf

                amt = (r.get("amount_cy") or r.get("amount") or 0)

                rows.append({
                    "entity_key":       entity_key,
                    "generic_key":      generic_key,
                    "dup_id":           "",
                    "Row Type":         "Original",
                    "Company Name":     entity,
                    "Statement":        stmt_kind,
                    "QB Account":       lbl,
                    "Parent path":      bc,
                    "Total $":          amt,
                    "abs_total":        abs(amt),
                    "Confidence":       confidence,
                    "Auto Suggestion":  auto or "",
                    "Target Line":      effective,
                    "Formula Override": pivot_override,
                    "Duplicate?":       False,
                    "Remove?":          False,
                })

                for de in dup_entries:
                    rows.append({
                        "entity_key":       entity_key,
                        "generic_key":      generic_key,
                        "dup_id":           de["dup_id"],
                        "Row Type":         "🧬 DUPLICATE",
                        "Company Name":     entity,
                        "Statement":        stmt_kind,
                        "QB Account":       lbl,
                        "Parent path":      bc,
                        "Total $":          amt,
                        "abs_total":        abs(amt),
                        "Confidence":       "duplicate",
                        "Auto Suggestion":  auto or "",
                        "Target Line":      de.get("target_line", ""),
                        "Formula Override": de.get("pivot_override", ""),
                        "Duplicate?":       False,
                        "Remove?":          False,
                    })
    rows.sort(key=lambda r: -r["abs_total"])
    return rows


rows = make_rows()
df   = pd.DataFrame(rows)

# Filters (explicit keys so selections survive reruns triggered by other
# buttons on this page, e.g. Apply duplicate/remove or Save Mappings)
chip_c1, chip_c2, chip_c3, chip_c4 = st.columns([2, 2, 2, 4])
with chip_c1:
    filter_stmt = st.radio("Statement", ["All", "P&L", "BS"],
                           horizontal=True, label_visibility="collapsed", key="v2_filter_stmt")
with chip_c2:
    filter_conf = st.radio("Confidence", ["All", "Need review", "Saved"],
                           horizontal=True, label_visibility="collapsed", key="v2_filter_conf")
with chip_c3:
    companies = ["All"] + sorted(qb_data.keys())
    filter_co = st.selectbox("Company", companies, label_visibility="collapsed", key="v2_filter_co")
with chip_c4:
    search = st.text_input("Search", "", label_visibility="collapsed",
                            placeholder="🔍 Search account name...", key="v2_filter_search")

view = df.copy()
if filter_stmt == "P&L":
    view = view[view["Statement"] == "P&L"]
elif filter_stmt == "BS":
    view = view[view["Statement"] == "BS"]
if filter_conf == "Need review":
    view = view[view["Confidence"] == "REVIEW"]
elif filter_conf == "Saved":
    view = view[view["Confidence"].isin(["saved", "entity-saved", "manual", "duplicate"])]
if filter_co != "All":
    view = view[view["Company Name"] == filter_co]
if search.strip():
    s = search.strip().lower()
    view = view[view["QB Account"].str.lower().str.contains(s, na=False) |
                view["Parent path"].str.lower().str.contains(s, na=False)]

view = view.drop(columns=["abs_total"], errors="ignore")

if target_lines_pnl or target_lines_bs:
    all_targets = sorted(set([""] + target_lines_pnl + target_lines_bs))
else:
    auto_suggestions = sorted(set(
        r["Auto Suggestion"] for _, r in df.iterrows()
        if r.get("Auto Suggestion") and r["Auto Suggestion"] != "__SKIP__"
    ))
    all_targets = sorted(set([""] + auto_suggestions))

edited = st.data_editor(
    view,
    width="stretch",
    hide_index=True,
    num_rows="fixed",
    column_order=[
        "Row Type", "Duplicate?", "Remove?",
        "Company Name", "Statement", "QB Account", "Parent path", "Total $",
        "Confidence", "Auto Suggestion", "Target Line", "Formula Override",
    ],
    column_config={
        "entity_key":     None,
        "generic_key":    None,
        "dup_id":         None,
        "Row Type":       st.column_config.TextColumn(
            "Row Type", disabled=True, width="small",
            help="🧬 DUPLICATE rows fan the same account out to an extra Target Line."),
        "Duplicate?":     st.column_config.CheckboxColumn(
            "🧬 Duplicate?", width="small", default=False,
            help="Check, then click 'Apply duplicate / remove' below to create a "
                 "blank copy of this row so you can map the same account to another "
                 "Target Line."),
        "Remove?":        st.column_config.CheckboxColumn(
            "🗑 Remove?", width="small", default=False,
            help="Check on a 🧬 DUPLICATE row, then click 'Apply duplicate / remove' "
                 "below to delete it. Has no effect on Original rows."),
        "Company Name":   st.column_config.TextColumn(disabled=True, width="medium"),
        "Statement":      st.column_config.TextColumn(disabled=True, width="small"),
        "QB Account":     st.column_config.TextColumn(disabled=True, width="medium"),
        "Parent path":    st.column_config.TextColumn(disabled=True, width="large"),
        "Total $":        st.column_config.NumberColumn(disabled=True, format="$%.0f", width="small"),
        "Confidence":     st.column_config.TextColumn(disabled=True, width="small"),
        "Auto Suggestion":st.column_config.TextColumn(disabled=True, width="medium"),
        "Target Line":    st.column_config.SelectboxColumn(
            "Target Line",
            options=all_targets if all_targets else [""],
            width="medium", required=False),
        "Formula Override": st.column_config.SelectboxColumn(
            "Formula Override",
            options=PIVOT_OVERRIDE_OPTIONS,
            width="medium", required=False,
            help="Pin this account's Target Line to a specific pivot tab "
                 "(P&L/BS Pivot CY, PY, or Change). Leave blank to use the "
                 "default sheet/auto-detected tab."),
    },
    height=600,
    key="variants_editor_main",
)


def _parse_entity_key(ekey: str):
    """Split 'E|stmt|entity|breadcrumb|label' -> (statement, entity, breadcrumb, label)."""
    parts = ekey.split("|", 4)
    if len(parts) != 5:
        return "", "", "", ""
    return parts[1], parts[2], parts[3], parts[4]


def _save_rows_to_profile(edited_df, base_overrides, entity_lookup_map, tid=None):
    """Merge Target Line / Formula Override edits from edited_df into a
    list-valued entity-mapping-overrides dict. If tid is given, also writes
    each touched row straight to MongoDB (profile `tid`) — this is the ONE
    place that path lives, used by Quick Save, the profile-picker dialog, and
    the Excel bulk-upload path, so they can never drift out of sync with
    each other.

    Returns (new_overrides_dict, n_saved, errors).
    """
    new_ov = dict(base_overrides)
    touched_keys = set()
    n = 0
    errors = []
    for _, r in edited_df.iterrows():
        ekey = r.get("entity_key") if hasattr(r, "get") else r["entity_key"]
        if not ekey:
            continue
        did      = (r.get("dup_id") or "") if hasattr(r, "get") else ""
        chosen   = (r.get("Target Line") or "").strip()
        pivot_ov = (r.get("Formula Override") or "").strip()

        if ekey not in touched_keys:
            # Seed with the current effective (DB + prior session) list so
            # rows hidden by the active filter (or absent from an uploaded
            # file) aren't dropped.
            new_ov[ekey] = [dict(e) for e in entity_lookup_map.get(ekey, [])]
            touched_keys.add(ekey)

        entries = new_ov[ekey]
        entries[:] = [e for e in entries if (e.get("dup_id") or "") != did]

        if chosen:
            entries.append({"dup_id": did, "target_line": chosen, "pivot_override": pivot_ov})
            n += 1
            if tid:
                stmt, ename, bc_v, lbl = _parse_entity_key(ekey)
                if ename:
                    try:
                        P.upsert_entity_mapping(
                            profile_id=tid, entity=ename, statement=stmt,
                            breadcrumb=bc_v, qb_account=lbl,
                            target_line=chosen, source="manual",
                            dup_id=did, pivot_override=pivot_ov,
                        )
                    except Exception as e:
                        errors.append(f"{ename} / {lbl}: {e}")
        elif tid:
            stmt, ename, bc_v, lbl = _parse_entity_key(ekey)
            if ename:
                try:
                    P.delete_entity_mapping(tid, ename, stmt, bc_v, lbl, dup_id=did)
                except Exception as e:
                    errors.append(f"{ename} / {lbl} (remove): {e}")
    return new_ov, n, errors


# ── Apply duplicate / remove — inline, immediate, no profile dialog ──────────
apply_col, apply_hint_col = st.columns([1, 3])
with apply_col:
    apply_clicked = st.button(
        "⚡ Apply duplicate / remove", width="stretch", key="v2_apply_dup_btn",
    )
with apply_hint_col:
    st.caption(
        "Applies immediately to this session (and deletes any checked "
        "duplicate from the profile right away). Whatever Target Line / "
        "Formula Override you've already typed above is kept — you still "
        "need **💾 Save Mappings** below to persist everything to the profile."
    )

if apply_clicked:
    new_session   = dict(st.session_state.get("entity_mapping_overrides", {}))
    touched_keys  = set()
    n_dup, n_rem  = 0, 0
    dup_errors    = []
    active_id_now = st.session_state.get("active_profile_id")

    for _, r in edited.iterrows():
        ekey     = r["entity_key"]
        did      = r.get("dup_id", "") or ""
        chosen   = (r["Target Line"] or "").strip()
        pivot_ov = (r.get("Formula Override", "") or "").strip()
        want_dup = bool(r.get("Duplicate?"))
        want_rem = bool(r.get("Remove?"))

        if ekey not in touched_keys:
            # Seed with the current effective (DB + prior session) list so
            # rows hidden by the active filter aren't dropped.
            new_session[ekey] = [dict(e) for e in entity_lookup.get(ekey, [])]
            touched_keys.add(ekey)

        entries = new_session[ekey]
        entries[:] = [e for e in entries if (e.get("dup_id") or "") != did]

        if want_rem and did:
            n_rem += 1
            if db_online and active_id_now:
                stmt_v, ename_v, bc_v, lbl_v = _parse_entity_key(ekey)
                if ename_v:
                    try:
                        P.delete_entity_mapping(active_id_now, ename_v, stmt_v, bc_v, lbl_v, dup_id=did)
                        P._invalidate_mapping_cache(active_id_now)
                    except Exception as e:
                        dup_errors.append(f"{ename_v} / {lbl_v}: {e}")
        elif did or chosen:
            # Keep the row (blank duplicates are kept too, so the user can
            # come back and fill in the Target Line before saving).
            entries.append({"dup_id": did, "target_line": chosen, "pivot_override": pivot_ov})

        if want_dup:
            n_dup += 1
            entries.append({"dup_id": uuid.uuid4().hex[:8], "target_line": "", "pivot_override": ""})

    st.session_state.entity_mapping_overrides = new_session

    if dup_errors:
        st.error("Some duplicates could not be removed from the profile: " + "; ".join(dup_errors))
    if n_dup or n_rem:
        st.success(
            f"✅ Applied: {n_dup} row(s) duplicated, {n_rem} duplicate(s) removed. "
            "New duplicate rows appear below with a blank Target Line — fill them "
            "in, then click 💾 Save Mappings to persist to the profile."
        )
        st.rerun()
    else:
        st.info("No 🧬 Duplicate or 🗑 Remove checkboxes were checked.")

# ── Quick Save — persists straight to the active profile, no dialog ─────────
st.divider()
qs_col, qs_hint_col = st.columns([1, 3])
_active_id_for_qs = st.session_state.get("active_profile_id")
with qs_col:
    quick_save_clicked = st.button(
        "⚡ Quick Save (active profile)", type="primary", width="stretch",
        key="v2_quick_save_btn", disabled=not (db_online and _active_id_for_qs),
    )
with qs_hint_col:
    if db_online and _active_id_for_qs:
        st.caption(
            f"Saves every Target Line / Formula Override above straight to profile "
            f"**{active_profile['name'] if active_profile else _active_id_for_qs}** — "
            "immediately, no profile picker. Use 💾 Save Mappings below instead if "
            "you want to switch or create a profile."
        )
    else:
        st.caption(
            "Pick or create a profile first (💾 Save Mappings below) to enable Quick Save."
        )

if quick_save_clicked:
    new_ov, n_saved, save_errors = _save_rows_to_profile(
        edited, entity_session_overrides, entity_lookup, tid=_active_id_for_qs,
    )
    st.session_state.entity_mapping_overrides = new_ov
    if save_errors:
        st.error(
            "⚠️ Saved to session, but some database writes failed:\n\n"
            + "\n".join(f"- {e}" for e in save_errors)
        )
    else:
        st.success(f"✅ Quick-saved {n_saved} mappings directly to the active profile.")
    st.rerun()

# ── Bulk edit via Excel — download the table, edit offline, re-upload ───────
st.divider()
st.markdown("##### 📥 Bulk edit via Excel")
st.caption(
    "Download the table currently shown above as a real .xlsx workbook, edit "
    "**Target Line** / **Formula Override** in Excel, then upload it back to "
    "replace those columns in this session in one shot. (Streamlit's own small "
    "download icon in the table's toolbar only exports CSV — use the button "
    "below instead for an Excel file.)"
)

_EXPORT_COLUMNS = [
    "Row Type", "Company Name", "Statement", "QB Account", "Parent path",
    "Total $", "Confidence", "Auto Suggestion", "Target Line", "Formula Override",
    "entity_key", "dup_id",
]


@st.cache_data(show_spinner=False, max_entries=4)
def _cached_mapping_excel(rows_hash: str, targets_hash: str,
                           _export_df: pd.DataFrame,
                           _pnl_targets: list, _bs_targets: list) -> bytes:
    """Write the mapping table to an .xlsx as plain values, with real Excel
    dropdowns so bulk-editing offline matches the in-app experience instead of
    requiring the whole Target Line to be typed out by hand:
      - Target Line: P&L rows get only P&L template lines as options, BS rows
        get only BS lines (via two per-row-scoped data validations).
      - Formula Override: the same fixed pivot-tab list for every row.
    Cached by content hash so unrelated reruns (e.g. a filter chip click
    elsewhere) don't pay to regenerate the file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapping"
    ws.append(_EXPORT_COLUMNS)
    for row in _export_df[_EXPORT_COLUMNS].itertuples(index=False, name=None):
        ws.append(list(row))
    n_rows = len(_export_df)

    # Hidden helper sheet holding the option lists — Excel's "list" data
    # validation needs a cell range for anything longer than a short inline set.
    pnl_opts = [""] + sorted({t for t in _pnl_targets if t})
    bs_opts  = [""] + sorted({t for t in _bs_targets if t})
    lists_ws = wb.create_sheet("Lists")
    lists_ws.sheet_state = "hidden"
    for i, v in enumerate(pnl_opts, start=1):
        lists_ws.cell(row=i, column=1, value=v)
    for i, v in enumerate(bs_opts, start=1):
        lists_ws.cell(row=i, column=2, value=v)
    for i, v in enumerate(PIVOT_OVERRIDE_OPTIONS, start=1):
        lists_ws.cell(row=i, column=3, value=v)

    target_col  = get_column_letter(_EXPORT_COLUMNS.index("Target Line") + 1)
    formula_col = get_column_letter(_EXPORT_COLUMNS.index("Formula Override") + 1)

    # Formula Override options are the same for every row — one range covers all.
    dv_formula = DataValidation(
        type="list", formula1=f"=Lists!$C$1:$C${len(PIVOT_OVERRIDE_OPTIONS)}", allow_blank=True,
    )
    ws.add_data_validation(dv_formula)
    if n_rows:
        dv_formula.add(f"{formula_col}2:{formula_col}{n_rows + 1}")

    # Target Line options depend on each row's Statement, so build them per row.
    dv_pnl = DataValidation(type="list", formula1=f"=Lists!$A$1:$A${len(pnl_opts)}", allow_blank=True) \
        if len(pnl_opts) > 1 else None
    dv_bs = DataValidation(type="list", formula1=f"=Lists!$B$1:$B${len(bs_opts)}", allow_blank=True) \
        if len(bs_opts) > 1 else None
    if dv_pnl is not None:
        ws.add_data_validation(dv_pnl)
    if dv_bs is not None:
        ws.add_data_validation(dv_bs)

    for i, stmt in enumerate(_export_df["Statement"], start=2):
        if stmt == "P&L" and dv_pnl is not None:
            dv_pnl.add(f"{target_col}{i}")
        elif stmt == "BS" and dv_bs is not None:
            dv_bs.add(f"{target_col}{i}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _rows_hash(export_df: pd.DataFrame) -> str:
    """Fast vectorized row hash (pandas' own C-accelerated hasher) — much
    cheaper than json.dumps for a wide table."""
    return hashlib.md5(
        pd.util.hash_pandas_object(export_df[_EXPORT_COLUMNS], index=False).values.tobytes()
    ).hexdigest()


def _targets_hash(pnl_targets: list, bs_targets: list) -> str:
    return hashlib.md5(
        ("|".join(sorted(pnl_targets)) + "##" + "|".join(sorted(bs_targets))).encode()
    ).hexdigest()


exp_col1, exp_col2 = st.columns([1, 3])
with exp_col1:
    st.download_button(
        "⬇ Download Excel",
        _cached_mapping_excel(
            _rows_hash(edited), _targets_hash(target_lines_pnl, target_lines_bs),
            edited, target_lines_pnl, target_lines_bs,
        ),
        "mapping_table.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
        key="v2_export_excel_btn",
    )

uploaded_mapping_file = st.file_uploader(
    "Upload edited mapping_table.xlsx",
    type=["xlsx"], key="v2_mapping_upload",
)
if uploaded_mapping_file is not None:
    if st.button("🔁 Replace table from upload", type="primary", key="v2_apply_upload_btn"):
        try:
            up_df = pd.read_excel(uploaded_mapping_file, dtype=str, engine="openpyxl").fillna("")
        except Exception as e:
            st.error(f"Could not read that file: {e}")
        else:
            missing = {"entity_key", "dup_id", "Target Line", "Formula Override"} - set(up_df.columns)
            if missing:
                st.error(
                    "That file is missing required columns: "
                    f"{', '.join(sorted(missing))}. Upload the file exactly as "
                    "downloaded above, only editing Target Line / Formula Override."
                )
            else:
                new_ov, n_up, up_errors = _save_rows_to_profile(
                    up_df, entity_session_overrides, entity_lookup, tid=None,
                )
                st.session_state.entity_mapping_overrides = new_ov
                st.success(
                    f"✅ Replaced {n_up} row(s) from the uploaded file into this session. "
                    "Review below, then click ⚡ Quick Save or 💾 Save Mappings to persist "
                    "to the profile and include it in the generated workbooks."
                )
                st.rerun()

# ── Save — dialog appears every time ──────────────────────────────────────────
st.divider()

if "v2_save_dialog_open" not in st.session_state:
    st.session_state.v2_save_dialog_open = False

save_col, hint_col = st.columns([1, 3])
with save_col:
    if st.button("💾 Save Mappings", type="primary",
                 width="stretch", key="v2_save_btn"):
        st.session_state.v2_save_dialog_open = True
with hint_col:
    if db_online:
        st.caption("You will be asked **which profile** to save to every time.")
    else:
        st.warning("MongoDB offline — mappings saved to session only.")

# ── Profile chooser dialog ─────────────────────────────────────────────────────
if st.session_state.get("v2_save_dialog_open"):
    st.markdown("---")
    with st.container(border=True):
        st.markdown("#### 📂 Save mappings to which profile?")

        if not db_online:
            st.warning("MongoDB is offline. Mappings will be saved to this browser session only.")
            oc1, oc2 = st.columns([1, 1])
            with oc1:
                if st.button("✅ Save to session", type="primary",
                             width="stretch", key="v2_sess_confirm"):
                    new_ov, n, _errs = _save_rows_to_profile(
                        edited, entity_session_overrides, entity_lookup, tid=None,
                    )
                    st.session_state.entity_mapping_overrides = new_ov
                    st.session_state.v2_save_dialog_open = False
                    st.success(f"✅ Saved {n} overrides to session.")
                    st.rerun()
            with oc2:
                if st.button("❌ Cancel", width="stretch", key="v2_sess_cancel"):
                    st.session_state.v2_save_dialog_open = False
                    st.rerun()
        else:
            _profiles_now = P.list_profiles()
            CREATE_NEW    = "__CREATE_NEW__"
            _pids   = [str(p["_id"]) for p in _profiles_now] + [CREATE_NEW]
            _plbls  = [p["name"]     for p in _profiles_now] + ["➕ Create new profile…"]

            _cur = st.session_state.get("active_profile_id", "")
            _def = _pids.index(_cur) if _cur in _pids else 0

            sel_id = st.selectbox(
                "Select profile to save into:",
                options=_pids,
                format_func=lambda x: _plbls[_pids.index(x)],
                index=_def,
                key="v2_dialog_sel",
            )

            new_name = ""
            if sel_id == CREATE_NEW:
                new_name = st.text_input(
                    "New profile name",
                    placeholder="Enter a name for the new profile…",
                    key="v2_dialog_new_name",
                )

            bc1, bc2 = st.columns([1, 1])
            with bc1:
                confirm = st.button("✅ Confirm Save", type="primary",
                                    width="stretch", key="v2_confirm")
            with bc2:
                if st.button("❌ Cancel", width="stretch", key="v2_cancel"):
                    st.session_state.v2_save_dialog_open = False
                    st.rerun()

            if confirm:
                if sel_id == CREATE_NEW:
                    if not new_name.strip():
                        st.warning("⚠️ Enter a name for the new profile first.")
                        st.stop()
                    tid   = str(P.create_profile(new_name.strip()))
                    tname = new_name.strip()
                else:
                    tid   = sel_id
                    tname = _plbls[_pids.index(sel_id)]

                st.session_state.active_profile_id = tid

                new_ov, n, save_errors = _save_rows_to_profile(
                    edited, entity_session_overrides, entity_lookup, tid=tid,
                )

                # Persist to session regardless of any per-row DB failures above,
                # so a single bad write can't silently wipe out everything else
                # the user just edited.
                st.session_state.entity_mapping_overrides = new_ov

                # Save row-level pivot overrides
                if db_online:
                    try:
                        d = dblib.get_db()
                        if d is not None:
                            d.mappings.delete_many({
                                "profile_id": P.ObjectId(tid),
                                "statement": "__template_row_override__"
                            })
                            for rk, val in st.session_state.get("row_pivot_overrides", {}).items():
                                if val:
                                    parts = rk.split("|", 1)
                                    if len(parts) == 2:
                                        sheet_name, row_idx = parts[0], parts[1]
                                        P.upsert_mapping(
                                            profile_id=tid,
                                            statement="__template_row_override__",
                                            breadcrumb=sheet_name,
                                            qb_account=row_idx,
                                            target_line=val,
                                            source="manual",
                                        )
                            P._invalidate_mapping_cache(tid)
                    except Exception as e:
                        save_errors.append(f"row formula overrides: {e}")
                st.session_state.loaded_row_overrides_profile_id = tid

                st.session_state.v2_save_dialog_open = False
                if save_errors:
                    st.error(
                        "⚠️ Saved to session, but some database writes failed — these "
                        "will NOT survive a page reload until fixed:\n\n"
                        + "\n".join(f"- {e}" for e in save_errors)
                    )
                else:
                    st.success(f"✅ Saved {n} per-company mappings and template row formula overrides to profile **{tname}**.")
                st.rerun()

st.divider()
if template_loaded:
    st.info("👉 Next: **💾 Generate Linked Workbook**.")
else:
    st.warning("Upload a target template on **📂 Upload Files** before generating.")
