"""Template-agnostic SUMIFS-linked workbook builder.

The builder reads the structure of the target template via `template_discovery`
and writes SUMIFS only into data-role cells. It preserves:
  • subtotal rows (formulas that sum across other rows)
  • cross-sheet references (equity rollforward chain)
  • preloaded rows (cells already containing static numbers — by default skipped)

For multi-year templates (BS 2025 / BS 2024 / IS 2025 / IS 2024 / …), each sheet's
SUMIFS pulls from a `QB_Data` sheet filtered by year and statement.

How rows are matched between QB accounts and template line items:
  1. Profile mappings (saved overrides) win.
  2. Auto-rules (mapping_rules.map_pnl / map_bs) fill remaining.
  3. Each mapped account → target_line. The template's column-A label IS used as
     the SUMIFS lookup key. If a template label changes spelling, the user
     re-runs and the engine fuzzy-matches old target_lines to new template
     labels via rapidfuzz (or simple normalize-and-equal if rapidfuzz absent).
"""

import io
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .mapping_rules import resolve_target_entries
from .master_builder import write_pivot_sheets
from .template_discovery import (
    discover_template, summarize, norm_label, YearSheet, RowInfo,
)

try:
    # pyrefly: ignore [missing-import]
    from rapidfuzz import fuzz, process as rf_process
    HAVE_RAPIDFUZZ = True
except ImportError:
    HAVE_RAPIDFUZZ = False


HDR_FILL = PatternFill("solid", fgColor="1F3864")
WHITE = Font(color="FFFFFF", bold=True)


def style_header(ws, row=1):
    for c in ws[row]:
        c.font = WHITE
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fmt_money(c):
    c.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def fit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


SYNONYMS = {
    # ── INCOME ───────────────────────────────────────────────────────────────
    "sales": ["sales", "retail sales", "food sales", "total sales", "ordinary income", "sales :-"],
    "retail sales": ["sales", "retail sales", "food sales", "total sales", "ordinary income", "sales :-"],
    "food sales": ["sales", "retail sales", "food sales", "total sales", "ordinary income", "sales :-"],
    "sales :-": ["sales", "retail sales", "food sales", "total sales", "ordinary income", "sales :-"],
    "rebate income": ["rebate income", "rebates", "discount received", "discounts received"],
    "discount received": ["rebate income", "rebates", "discount received", "discounts received"],
    "cam charges collected": ["cam charges collected", "cam income", "cam charges income"],
    "income from partnership / k-1": ["income from partnership / k-1", "k1 income", "k-1 income", "income from k1", "income from k-1", "income from partnership"],
    "income from other entities": ["income from other entities", "income from other", "other entity income"],
    "management fees income": ["management fees  income", "management fees income", "management fee income"],
    "erc credit income": ["erc credit income", "erc income", "employee retention credit"],
    "eidl grant income": ["eidl grant income", "eidl grant", "eidl stimulus"],
    "interest income": ["interest income", "interest earned", "bank interest"],
    "rental income": ["rental income", "rent income"],
    "non taxable ppp income": ["non taxable ppp income", "ppp income", "ppp loan forgiveness"],

    # ── COGS / LABOR ─────────────────────────────────────────────────────────
    "salaries & wages": ["salaries & wages", "regular hours", "salary", "ot hours", "salaries and wages", "salary & wages", "wages", "salaries", "salaries wages"],
    "regular hours": ["regular hours", "salaries & wages", "salaries and wages"],
    "ot hours": ["ot hours", "overtime", "salaries & wages"],
    "kitchen labor": ["kitchen labor", "kitchen labour"],
    "training cost": ["training cost", "training", "training expense"],
    "payroll taxes": ["payroll taxes", "payroll tax", "fica", "social security tax", "unemployment tax"],
    "bonus": ["bonus", "bonus expsnse", "bonus expense"],
    "performance bonus": ["performance bonus", "performance based bonus"],
    "profit sharing expense": ["profit sharing expense", "profit sharing", "profit transfer"],
    "covid care": ["covid care", "covid pay", "covid sick pay"],
    "insurance & workers' compensation": ["insurance & workers' compensation", "insurance exp & workers comp", "insurance & workers comp", "insurance exp", "insurance expense", "insurance", "workers compensation", "workers comp"],

    # ── COGS / FRANCHISE ─────────────────────────────────────────────────────
    "franchise & advt fees": ["franchise & advt fees", "total franchise fees", "franchise fees", "franchise fee", "royalty", "ad fund"],
    "franchise preservation fees": ["franchise preservation fees", "franchise preservation", "preservation fee"],

    # ── COGS / PURCHASE ──────────────────────────────────────────────────────
    "purchase": ["purchase", "purchases", "food purchase", "food purchases"],
    "opening inventory": ["opening inventory", "beginning inventory", "inventory at 01/01"],
    "ending inventory": ["ending inventory", "closing inventory", "inventory at 12/31", "less: inventory"],

    # ── COGS / DELIVERY ──────────────────────────────────────────────────────
    "delivery fees expense": ["delivery fees expense", "delivery charges", "delivery expense", "doordash", "grubhub", "ubereats", "uber eats", "otter"],
    "delivery charges": ["delivery charges", "delivery fees expense", "delivery expense"],
    "delivery sales receivable": ["delivery sales receivable", "receivable from deliveries", "delivery receivable", "delivery sales  receivable"],

    # ── COGS / OTHER ─────────────────────────────────────────────────────────
    "restaurant supplies": ["restaurant supplies", "restaurant supp", "restaurant supply", "kitchen smallware", "smallware"],
    "restaurant supp": ["restaurant supp", "restaurant supplies", "restaurant supply"],
    "uniforms": ["uniforms", "uniform", "uniform "],
    "uniform": ["uniform", "uniforms", "uniform "],
    "gift card charges": ["gift card charges", "gift card expense"],
    "food for employees": ["food for employees", "employee meals", "employee food"],

    # ── DIGITAL / POPEYES ────────────────────────────────────────────────────
    "popeyes digital transaction fee": ["popeyes digital transaction fee", "digital transaction fee", "digital transaction"],
    "popeyes ordering technology fee": ["popeyes ordering technology fee", "ordering technology fee", "ordering tech fee"],
    "popeyes guest care fee": ["popeyes guest care fee", "guest care fee", "guest care"],
    "popeyes service check": ["popeyes service check", "service check", "popeyes service"],

    # ── OPERATING EXPENSES ───────────────────────────────────────────────────
    "auto expenses and travel": ["auto expenses and travel", "auto expense", "auto expenses", "auto expense & travel", "auto and travel"],
    "auto expense": ["auto expense", "auto expenses and travel", "auto expenses"],
    "advertising and promotion": ["advertising and promotion", "advertising", "advertise", "promotion", "marketing"],
    "bank service charges": ["bank service charges", "bank charges", "bank service", "bank fees"],
    "cleaning expenses": ["cleaning expenses", "cleaning exp", "cleaning expense", "cleaning"],
    "rent & cam charges": ["rent & cam charges", "rent", "rent expense", "rent & cam", "cam charges"],
    "utilities": ["utilities", "utility", "utilities expense", "electricity", "water", "gas"],
    "real estate tax": ["real estate tax", "re tax and proprty tax", "property tax", "real estate  tax"],
    "repairs and maintenance": ["repairs and maintenance", "repair & maint", "repair and maintenance", "repairs & maintenance", "repairs", "maintenance"],
    "credit card charges": ["credit card charges", "merchant fees", "merchant card charges", "credit card fee"],
    "management fees": ["management fees", " management fees", "    management fees", "mgmt fees"],
    "legal fees": ["legal fees", "legal expense", "attorney fees"],
    "professional fees": ["professional fees", "professional expense", "accounting fees", "accounting"],
    "alarm and security": ["alarm and security", "security and alarm", "alarm & security", "security & alarm", "security expense"],
    "licenses and permits": ["licenses and permits", "licence and permits", "license and permits", "licenses & permits", "license and permit"],
    "office supplies and expense": ["office supplies and expense", "office expenses", "office supplies", "office expense"],
    "payroll processing": ["payroll processing", "payroll expense", "payroll service fees", "payroll processing fees"],
    "401(k) expenses": ["401(k) expenses", "401k expense", "401k expenses", "payroll expense"],
    "payroll expense": ["payroll expense", "payroll processing", "401(k) expenses"],
    "shortage and overs": ["shortage and overs", "shortages", "shortages & overs", "shortage & over", "cash short/over"],
    "shortages": ["shortages", "shortages & overs", "shortage & over", "cash short/over", "shortage and overs"],
    "kiosk fees": ["kiosk fees", "kiosk fee", "kiosk"],
    "dues and subscriptions": ["dues and subscriptions", "dues", "subscriptions", "subscription"],
    "cash handling service": ["cash handling service", "cash handling"],
    "house charge": ["house charge", "house charges"],
    "convention expense": ["convention expense", "convention expenses", "convention", "convension expenses"],
    "refinance charges": ["refinance charges", "refinancing charges"],
    "guaranteed payment": ["guaranteed payment", "guaranteed payments"],
    "brokerage fees": ["brokerage fees", "brokerage fee", "brokerage"],
    "software expense": ["software expense", "software", "computer expenses", "computer expense"],
    "equipment rental": ["equipment rental", "equipment lease"],
    "donation": ["donation", "donations", "charity"],
    "loan fees": ["loan fees", "loan fee"],
    "miscellaneous expenses": ["miscellaneous expenses", "misc exp", "miscellaneous expense", "misc expense", "misc"],
    "annual report fees": ["annual report fees", "annual report fee", "annual report"],
    "state filing fees": ["state filing fees", "state filing fees payable", "annual report fees", "filing fees", "filing fee", "nj filing"],
    "state taxes / illinois replacement tax": ["state taxes / illinois replacement tax", "state taxes", "illinois replacement tax", "state tax"],
    "meal taxes": ["meal taxes", "meal tax", "restaurant tax", "misc taxes"],
    "penalty and interest": ["penalty and interest", "penalty", "interest penalty"],
    "interest to bank": ["interest to bank", "bank interest expense", "interest expense bank"],
    "interest to others": ["interest to others", "interest expense", "interest to other"],
    "depreciation": ["depreciation", "depreciation expense"],
    "amortization": ["amortization", "amortisation"],

    # ── NON-CASH / ADJUSTMENTS ────────────────────────────────────────────────
    "gain / (loss) on sale of assets": ["gain / (loss) on sale of assets", "gain on sale", "loss on sale", "gain loss on disposal"],
    "non-deductible expense": ["non-deductible expense", "non deductible expense"],
    "wotc non-taxable expense": ["wotc non-taxable expense", "wotc", "wotc credit"],
    "development rights written off": ["development rights written off", "development right write off"],
    "donation from k-1 / passthrough": ["donation from k-1 / passthrough", "donation from k1", "passthrough donation"],
    "1231 loss from k-1": ["1231 loss from k-1", "1231 loss", "section 1231 loss"],
    "profit / loss transfer to management entity": ["profit / loss transfer to management entity", "profit transfer to management", "profit transfer"],
    "profit transfer to ap north": ["profit transfer to ap north", "profit sharing expense", "profit sharing"],

    # ── BALANCE SHEET — ASSETS ────────────────────────────────────────────────
    "cash on hand": ["cash on hand", "petty cash", "cash on hand and in bank"],
    "cash in bank": ["cash in bank", "cash on hand and in bank", "pnc checking", "bank account", "checking account"],
    "cash on hand and in bank": ["cash on hand and in bank", "cash on hand", "cash in bank", "cash on hand & in bank", "total cash & bank"],
    "credit card receivable": ["credit card receivable", "credit card receivables"],
    "due from (to) affiliates": ["due from (to) affiliates", "due from/(due to) affiliates", "due from/to affiliates", "due to from affiliates"],
    "inventory": ["inventory", "ending inventory", "opening inventory", "food inventory"],
    "loan receivable from partners": ["loan receivable from partners", "loan receivable", "loan recievable", "due from partners"],
    "investment in other business": ["investment in other business", "investment in business"],
    "investment in affiliates": ["investment in affiliates", "investment in affiliate"],
    "security deposit": ["security deposit", "security deposits"],
    "escrow deposit": ["escrow deposit", "escrow"],
    "work in progress": ["work in progress", "wip", "construction in progress"],
    "exchanges": ["exchanges", "exchange", "1031 exchange"],

    # ── BALANCE SHEET — FIXED / INTANGIBLE ───────────────────────────────────
    "equipments": ["equipments", "equipment", "furniture fixtures", "ffe"],
    "less: accumulated depreciation": ["less: accumulated depreciation", "accumulated depreciation", "acc depreciation"],
    "goodwill": ["goodwill"],
    "franchise fees": ["franchise fees", "franchise fee", "franchise cost"],
    "leasehold imp. (intangible)": ["leasehold imp. (intangible)", "leasehold improvements", "leasehold improvemrnts", "leasehold improvement", "leasehold imp"],
    "organization costs": ["organization costs", "organization expenses", "organizational costs"],
    "deferred financing costs": ["deferred financing costs", "loan cost", "deferred loan cost", "financing costs"],
    "closing costs": ["closing costs", "closing cost"],
    "development rights": ["development rights", "development right"],

    # ── BALANCE SHEET — LIABILITIES ──────────────────────────────────────────
    "accounts payable": ["accounts payable", "ap"],
    "accrued expenses": ["accrued expenses", "accrued liabilities", "accrued expense"],
    "sales tax payable": ["sales tax payable", "sales tax"],
    "payroll liabilities": ["payroll liabilities", "payroll taxes payable", "payroll tax payable"],
    "state tax payable": ["state tax payable", "state tax payable-maryland", "state income tax payable"],
    "non-resident tax payable": ["non-resident tax payable", "non resident tax payable", "md non resident tax"],
    "net payroll checks payable": ["net payroll checks payable", "net payroll", "payroll checks payable"],
    "meal taxes payable": ["meal taxes payable", "meal tax payable"],
    "security deposit payable": ["security deposit payable"],
    "nj filing fees payable": ["nj filing fees payable", "nj annual filing fees payable"],
    "insurance proceeds payable": ["insurance proceeds payable"],
    "plk donation payable": ["plk donation payable", "plk donation"],
    "tb foundation payable": ["tb foundation payable", "tb foundation"],
    "loan payable - others (current)": ["loan payable - others (current)", "loan payable others current"],
    "other current liabilities": ["other current liabilities", "other current liab"],
    "loan payable to bank": ["loan payable to bank", "bank loan payable", "loan payable bank"],
    "loan payable to sba / eidl": ["loan payable to sba / eidl", "eidl loan payable", "sba loan", "eidl"],
    "ppp loan payable": ["ppp loan payable", "ppp loan", "ppp payable"],
    "loan payable to partners": ["loan payable to partners", "loan payable partner", "partner loan payable"],
    "loan payable to others": ["loan payable to others", "loan payable other"],

    # ── BALANCE SHEET — EQUITY ────────────────────────────────────────────────
    "beginning capital (partners' capital - beginning)": ["beginning capital (partners' capital - beginning)", "partner's capital-beging", "beginning capital", "partner capital beginning", "opening capital"],
    "retained earnings": ["retained earnings", "retained earning"],
    "add: current year profit / (loss)": ["add: current year profit / (loss)", "current year profit", "net profit", "net income"],
    "distribution paid to partner": ["distribution paid to partner", "partner distribution", "distributions"],
    "local mercantile tax": ["local mercantile tax", "mercantile tax", "bpt", "school tax"],
    "state filing fees payable": ["state filing fees payable", "state filing fees", "annual report fees", "filing fees"],
}

# Pre-normalize the synonyms at runtime once
NORM_SYNONYMS = {}
for k, v in SYNONYMS.items():
    kn = norm_label(k)
    if not kn:
        continue
    vn_list = NORM_SYNONYMS.setdefault(kn, [])
    for val in v:
        vn = norm_label(val)
        if vn and vn not in vn_list:
            vn_list.append(vn)


def _best_template_row_for(target_line: str, candidates: List[str], threshold: int = 70) -> Optional[str]:
    """Given a target_line (from the mapping rules), find the closest template
    label among the candidates."""
    if not target_line or not candidates:
        return None

    tnorm = norm_label(target_line)
    if not tnorm:
        return None

    cand_norm = {norm_label(c): c for c in candidates}

    # 1. Exact normalized match
    if tnorm in cand_norm:
        return cand_norm[tnorm]

    # 2. Hardcoded synonyms / CPA common equivalents
    if tnorm in NORM_SYNONYMS:
        for syn_norm in NORM_SYNONYMS[tnorm]:
            if syn_norm in cand_norm:
                return cand_norm[syn_norm]

    for cand_n, original_cand in cand_norm.items():
        if cand_n in NORM_SYNONYMS:
            if tnorm in NORM_SYNONYMS[cand_n]:
                return original_cand

    # 3. Substring matching (if one contains the other)
    for cand_n, original_cand in cand_norm.items():
        if len(cand_n) > 3 and len(tnorm) > 3:
            if cand_n in tnorm or tnorm in cand_n:
                return original_cand

    # 4. Standard library difflib SequenceMatcher (always available fallback)
    import difflib
    best_ratio = 0
    best_cand = None
    for cand_n, original_cand in cand_norm.items():
        ratio = difflib.SequenceMatcher(None, tnorm, cand_n).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_cand = original_cand
    if best_ratio >= 0.75:
        return best_cand

    # 5. Rapidfuzz if installed
    if HAVE_RAPIDFUZZ:
        best = rf_process.extractOne(tnorm, list(cand_norm.keys()),
                                      scorer=fuzz.token_set_ratio,
                                      score_cutoff=threshold)
        if best:
            return cand_norm[best[0]]

    return None


def compute_mapping(qb_data, mapping_overrides=None, entity_mapping_overrides=None):
    """Run auto-rules on every leaf QB account, allow overrides. Returns:
        mapping: dict[(statement, entity, breadcrumb, qb_account)] =
                 list[(target_line, source, pivot_override)]
        pnl_leaves, bs_leaves: dict[(breadcrumb, label)] = {amounts: {entity: $}}

    Lookup priority per entity+account (see mapping_rules.resolve_target_entries):
      1. Entity-specific override  (key: 'E|stmt|entity|bc|lbl')
      2. Generic override          (key: 'stmt|bc|lbl')
      3. Auto-rules
      4. Any "duplicate" entity-level mappings, appended as extra list entries
         so one account can fan out into more than one Target Line.
    """
    overrides        = mapping_overrides or {}
    entity_overrides = entity_mapping_overrides or {}
    pnl_leaves: Dict = {}
    bs_leaves:  Dict = {}
    mapping:    Dict = {}

    for entity, info in qb_data.items():
        for r in info.get("pnl_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            k = (r["breadcrumb"], r["label"])
            d = pnl_leaves.setdefault(k, {"amounts_cy": {}, "amounts_py": {},
                                          "breadcrumb": r["breadcrumb"], "label": r["label"]})
            d["amounts_cy"][entity] = d["amounts_cy"].get(entity, 0) + (r.get("amount_cy") or r.get("amount") or 0)
            if r.get("amount_py") is not None:
                d["amounts_py"][entity] = d["amounts_py"].get(entity, 0) + (r.get("amount_py") or 0)
        for r in info.get("bs_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            k = (r["breadcrumb"], r["label"])
            d = bs_leaves.setdefault(k, {"amounts_cy": {}, "amounts_py": {},
                                         "breadcrumb": r["breadcrumb"], "label": r["label"]})
            d["amounts_cy"][entity] = d["amounts_cy"].get(entity, 0) + (r.get("amount_cy") or r.get("amount") or 0)
            if r.get("amount_py") is not None:
                d["amounts_py"][entity] = d["amounts_py"].get(entity, 0) + (r.get("amount_py") or 0)

    # Build per-entity mapping — entity-specific override wins
    for entity, info in qb_data.items():
        for r in info.get("pnl_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            bc, lbl = r["breadcrumb"], r["label"]
            mkey = ("P&L", entity, bc, lbl)
            mapping[mkey] = resolve_target_entries("P&L", entity, bc, lbl, overrides, entity_overrides)

        for r in info.get("bs_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            bc, lbl = r["breadcrumb"], r["label"]
            mkey = ("BS", entity, bc, lbl)
            mapping[mkey] = resolve_target_entries("BS", entity, bc, lbl, overrides, entity_overrides)

    return mapping, pnl_leaves, bs_leaves


def _write_qb_data_sheet(wb, qb_data, mapping):
    """Hidden source sheet: every leaf row × CY/PY × entity, with resolved target_line."""
    sheet_name = "QB_Data"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["Year", "Entity", "Statement", "Breadcrumb", "QB Account",
               "Target Line", "Mapping Key", "Amount"])
    style_header(ws)
    ws.freeze_panes = "A2"

    for entity, info in qb_data.items():
        # Write a "CY" row and (if available) a "PY" row for each leaf, per
        # resolved target_line (an account with duplicate mappings writes one
        # set of rows per target_line so each fans out correctly).
        for r in info.get("pnl_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            entries = mapping.get(("P&L", entity, r["breadcrumb"], r["label"])) or [("", "REVIEW", "")]
            key = f"P&L|{r['breadcrumb']}|{r['label']}"
            cy = r.get("amount_cy") or r.get("amount") or 0
            py = r.get("amount_py")
            for tgt, _src, _po in entries:
                tgt = (tgt or "").strip()
                ws.append(["CY", entity, "P&L", r["breadcrumb"], r["label"], tgt, key, cy])
                fmt_money(ws.cell(row=ws.max_row, column=8))
                if py is not None:
                    ws.append(["PY", entity, "P&L", r["breadcrumb"], r["label"], tgt, key, py])
                    fmt_money(ws.cell(row=ws.max_row, column=8))
        for r in info.get("bs_rows", []):
            if r["is_section_only"] or r["is_total"]:
                continue
            entries = mapping.get(("BS", entity, r["breadcrumb"], r["label"])) or [("", "REVIEW", "")]
            key = f"BS|{r['breadcrumb']}|{r['label']}"
            cy = r.get("amount_cy") or r.get("amount") or 0
            py = r.get("amount_py")
            for tgt, _src, _po in entries:
                tgt = (tgt or "").strip()
                ws.append(["CY", entity, "BS", r["breadcrumb"], r["label"], tgt, key, cy])
                fmt_money(ws.cell(row=ws.max_row, column=8))
                if py is not None:
                    ws.append(["PY", entity, "BS", r["breadcrumb"], r["label"], tgt, key, py])
                    fmt_money(ws.cell(row=ws.max_row, column=8))
    fit(ws, [6, 28, 6, 50, 35, 35, 60, 16])
    # Hide it from casual view
    ws.sheet_state = "hidden"


def _write_template_map_sheet(wb, year_sheets, mapping):
    """Audit sheet that records, per template year-sheet, every data-row label
    and which QB target_lines route to it. Useful for diagnosing why a cell is 0."""
    if "Template_Map" in wb.sheetnames:
        del wb["Template_Map"]
    ws = wb.create_sheet("Template_Map", 1)
    ws.append(["Sheet", "Row", "Template Label", "Matched Target Line", "Source QB Accounts"])
    style_header(ws)
    ws.freeze_panes = "A2"

    # Build reverse index: target_line → list of QB account keys
    by_target = {}
    for (stmt, entity, bc, lbl), entries in mapping.items():
        for tgt, src, po in entries:
            if tgt:
                by_target.setdefault(tgt, []).append((stmt, bc, lbl))

    for ys in year_sheets:
        for row in ys.rows:
            if row.role != "data":
                continue
            # Find which target_line(s) match this template label, filtered by statement type
            stmt_key = "P&L" if ys.statement == "IS" else ys.statement
            candidates = [tgt for tgt, qbs in by_target.items() if any(stmt == stmt_key for stmt, _, _ in qbs)]
            matched = _best_template_row_for(row.label, candidates)
            qbs = by_target.get(matched, []) if matched else []
            qb_summary = "; ".join(f"{stmt}: {lbl}" for stmt, bc, lbl in qbs[:5])
            if len(qbs) > 5:
                qb_summary += f"  … (+{len(qbs)-5} more)"
            ws.append([ys.sheet_name, row.row_idx, row.label, matched or "(no match)", qb_summary])
    fit(ws, [22, 6, 40, 40, 80])


def _resolve_row_to_target(template_label: str, target_lines: List[str]) -> Optional[str]:
    """Fuzzy-match a template label against known target_lines."""
    return _best_template_row_for(template_label, target_lines)


def build_linked_workbook(
    qb_data,
    target_bytes,
    mapping_overrides=None,
    entity_mapping_overrides=None,
    overwrite_preloaded: bool = True,
    selected_sheets=None,
    entity_col_mapping=None,
    sheet_pivot_overrides=None,
    row_pivot_overrides=None,
):
    """Build the linked combination workbook. Returns (BytesIO, mapping, year_sheets, report).

    Args:
        selected_sheets: Optional list of template sheet names to process.
            If None, all IS/BS-classified sheets are processed.
        entity_col_mapping: Optional dict mapping template column header name
            to the QB entity name whose data should flow into that column.
            A value of None/missing means use the column header verbatim
            (auto-matching by name). Set a column's value to None to skip it.
        sheet_pivot_overrides: Optional dict mapping template sheet name to
            "CY" or "PY" — overrides the auto-detected year filter used when
            building the pivot tab name in each SUMIFS formula.  E.g.:
            {"BS 2024": "PY", "IS 2025": "CY"}
        row_pivot_overrides: Optional dict mapping "sheet_name|row_idx" to the exact
            pivot sheet name to use for that row's SUMIFS formulas.
    """
    wb = openpyxl.load_workbook(io.BytesIO(target_bytes))
    year_sheets = discover_template(wb)

    # Filter to selected sheets only
    if selected_sheets:
        year_sheets = [ys for ys in year_sheets if ys.sheet_name in selected_sheets]

    if not year_sheets:
        raise RuntimeError(
            f"No IS/BS sheets found in the selection. "
            f"Sheets present: {wb.sheetnames}. "
            f"Selected: {selected_sheets}"
        )

    mapping, pnl_leaves, bs_leaves = compute_mapping(
        qb_data, mapping_overrides, entity_mapping_overrides
    )
    
    # Delete old pivot/QB_Data sheets if they exist in the uploaded template
    for suffix in ["CY", "PY", "Change"]:
        for base in ["P&L Pivot", "BS Pivot"]:
            sn = f"{base} {suffix}"
            if sn in wb.sheetnames:
                del wb[sn]
    if "QB_Data" in wb.sheetnames:
        del wb["QB_Data"]

    # Write fresh pivot sheets containing the Target Line column (Column C)
    pnl_pivot_full, bs_pivot_full = write_pivot_sheets(
        wb, qb_data, mapping_overrides, entity_mapping_overrides
    )
    _write_template_map_sheet(wb, year_sheets, mapping)

    # Per-Target-Line pivot-tab preference set via the main mapping table's
    # "Formula Override" column (either on the original row or a duplicate).
    # Lower priority than row_pivot_overrides (a specific template row), higher
    # than sheet_pivot_overrides/auto (see priority chain below).
    target_line_pivot_overrides: Dict[str, str] = {}
    for pv in (pnl_pivot_full, bs_pivot_full):
        for v in pv.values():
            if v["is_section"] or v["is_total"]:
                continue
            tgt = v.get("target_line")
            po = v.get("pivot_override")
            if tgt and po and tgt not in target_line_pivot_overrides:
                target_line_pivot_overrides[tgt] = po

    # Available target_lines (from mapping output)
    pnl_targets = sorted({t for (s, _, _, _), entries in mapping.items() if s == "P&L" for (t, _, _) in entries if t})
    bs_targets = sorted({t for (s, _, _, _), entries in mapping.items() if s == "BS" for (t, _, _) in entries if t})

    report = {
        "n_year_sheets": len(year_sheets),
        "cells_written": 0,
        "cells_skipped_preloaded": 0,
        "cells_skipped_subtotal": 0,
        "cells_skipped_crossref": 0,
        "rows_unmapped": [],
    }

    entities = list(qb_data.keys())

    for ys in year_sheets:
        ws = wb[ys.sheet_name]
        is_bs = ys.statement == "BS"
        year_filter = "CY" if (ys.year is None or _is_cy_year(ys, year_sheets)) else "PY"
        targets = bs_targets if is_bs else pnl_targets

        for row in ys.rows:
            if row.role == "subtotal":
                report["cells_skipped_subtotal"] += len(ys.entity_cols)
                continue
            if row.role == "cross_ref":
                report["cells_skipped_crossref"] += len(ys.entity_cols)
                continue
            if row.role == "preloaded" and not overwrite_preloaded:
                report["cells_skipped_preloaded"] += len(ys.entity_cols)
                continue
            if row.role not in ("data", "preloaded"):
                continue

            # First, clean/clear the existing values in all entity columns for this row
            for col, entity_header in ys.entity_cols:
                ws.cell(row=row.row_idx, column=col).value = None

            # Match this template label to a target_line
            matched_target = _resolve_row_to_target(row.label, targets)
            if not matched_target:
                report["rows_unmapped"].append({
                    "sheet": ys.sheet_name, "row": row.row_idx, "label": row.label,
                })
                continue

            # Write SUMIFS into each entity column
            for col, entity_header in ys.entity_cols:
                cell = ws.cell(row=row.row_idx, column=col)
                # If preloaded and not overwriting, skip per-cell (re-check)
                if row.role == "preloaded" and not overwrite_preloaded:
                    continue

                # Resolve the matching QB entity and year filter using the smart resolver
                qb_entity, resolved_year_filter = resolve_qb_source(
                    ys, year_sheets, qb_data, entity_col_mapping, entity_header
                )
                
                if qb_entity is None:
                    # Skipped or not found: leave cell untouched
                    continue

                if qb_entity in entities:
                    idx = entities.index(qb_entity)
                    col_letter = get_column_letter(6 + idx)  # Column F onwards (6 = F)
                else:
                    # Entity not uploaded/found: leave cell untouched
                    continue

                # Priority: per-row override (Template Row Formula Overrides) >
                # per-Target-Line override (Formula Override column on the main
                # mapping table) > per-sheet override > auto-detected year filter.
                row_key = f"{ys.sheet_name}|{row.row_idx}"
                if row_pivot_overrides and row_pivot_overrides.get(row_key):
                    pivot_sheet_name = row_pivot_overrides[row_key]
                elif matched_target in target_line_pivot_overrides:
                    pivot_sheet_name = target_line_pivot_overrides[matched_target]
                else:
                    sheet_base = "BS Pivot" if ys.statement == "BS" else "P&L Pivot"
                    # Apply per-sheet pivot tab override if provided;
                    # otherwise fall back to the auto-detected year filter.
                    if sheet_pivot_overrides and ys.sheet_name in sheet_pivot_overrides:
                        effective_year_filter = sheet_pivot_overrides[ys.sheet_name]
                    else:
                        effective_year_filter = resolved_year_filter
                    pivot_sheet_name = f"{sheet_base} {effective_year_filter}"
                # Use the template's detected label column (not hardcoded $A).
                # AP Illinois: label_col=2 → $B; NJ/original: label_col=1 → $A
                label_col_letter = get_column_letter(ys.label_col)
                # TRIM() on both the lookup cell and the Target Line column (col C) in pivot.
                # This ensures Excel SUMIFS exact-match is never broken by leading/trailing spaces
                # in either the template labels or the auto-mapped target lines.
                formula = (
                    f"=SUMIFS('{pivot_sheet_name}'!${col_letter}:${col_letter}, "
                    f"'{pivot_sheet_name}'!$C:$C, "
                    f"TRIM(${label_col_letter}{row.row_idx}))"
                )

                cell.value = formula
                fmt_money(cell)
                report["cells_written"] += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, mapping, year_sheets, report


def _is_cy_year(ys: YearSheet, all_sheets: List[YearSheet]) -> bool:
    """The newest year among sheets of the same statement is the CY."""
    years = [s.year for s in all_sheets if s.statement == ys.statement and s.year]
    if not years:
        return True
    return ys.year == max(years)


def resolve_qb_source(
    ys: YearSheet,
    all_sheets: List[YearSheet],
    qb_data: dict,
    entity_col_mapping: Optional[dict],
    entity_header: str,
) -> Tuple[Optional[str], str]:
    """Resolves the correct QB entity key and the year filter ('CY' or 'PY') for a given template column and sheet.

    Returns:
        (resolved_qb_entity_key, year_filter)
    """
    is_cy = _is_cy_year(ys, all_sheets)

    # 1. Check manual mapping first
    mapped_entity = None
    if entity_col_mapping is not None:
        if entity_header in entity_col_mapping:
            mapped_entity = entity_col_mapping[entity_header]
            if mapped_entity is None:
                # Explicitly mapped to "Skip"
                return None, "CY"

    # 2. Extract metadata and clean names for all uploaded QB entities
    candidates = []
    for ent_key, info in qb_data.items():
        # Detect year from period strings or filename
        ent_year = None
        for p in [info.get("pnl_period_cy"), info.get("bs_period_cy"), info.get("pnl_period"), info.get("bs_period")]:
            if p:
                p_str = str(p).strip()
                m = re.search(r"\b(20\d{2})\b", p_str)
                if m:
                    ent_year = int(m.group(1))
                    break
                m2 = re.search(r"\b(\d{2})$", p_str)
                if m2:
                    val = int(m2.group(1))
                    if 0 <= val <= 99:
                        ent_year = 2000 + val if val < 50 else 1900 + val
                        break
                # Look for a 2 digit year with a preceding boundary, e.g. - 24, / 24, space 24
                m3 = re.search(r"\b(\d{2})\b", p_str)
                if m3:
                    val = int(m3.group(1))
                    if 15 <= val <= 35:
                        ent_year = 2000 + val
                        break
        
        if not ent_year:
            m_fn = re.search(r"\b(20\d{2})\b", info.get("file", ""))
            if m_fn:
                ent_year = int(m_fn.group(1))
            else:
                m_fn2 = re.search(r"\b(\d{2})\b", info.get("file", ""))
                if m_fn2:
                    val = int(m_fn2.group(1))
                    if 20 <= val <= 35:
                        ent_year = 2000 + val

        # Clean the name: normalize and strip any year numbers out
        clean_ent = norm_label(ent_key)
        clean_ent_no_year = clean_ent
        if ent_year:
            yr_str = str(ent_year)
            yr_short = str(ent_year % 100)
            clean_ent_no_year = re.sub(r"\b(" + yr_str + r"|" + yr_short + r")\b", "", clean_ent)
            clean_ent_no_year = re.sub(r"[\s\-_]+", " ", clean_ent_no_year).strip()

        candidates.append({
            "key": ent_key,
            "clean_name": clean_ent_no_year,
            "year": ent_year,
            "variant": info.get("variant", "single")
        })

    # If manual mapping is specified, find that entity key
    if mapped_entity:
        cand = next((c for c in candidates if c["key"] == mapped_entity), None)
        if cand:
            if cand["variant"] == "triple":
                return cand["key"], ("CY" if is_cy else "PY")
            else:
                return cand["key"], "CY"

    # Otherwise, auto-resolve by comparing normalized name with normalized header
    clean_header = norm_label(entity_header)
    
    # Try exact match on clean name first
    matches = [c for c in candidates if c["clean_name"] == clean_header or norm_label(c["key"]) == clean_header]
    # Fallback to substring matching if needed
    if not matches:
        matches = [c for c in candidates if clean_header in c["clean_name"] or clean_header in norm_label(c["key"])]

    if not matches:
        return None, "CY"

    # Case A: If one of the matching files is a multi-year (triple) file, use it
    triple_match = next((c for c in matches if c["variant"] == "triple"), None)
    if triple_match:
        return triple_match["key"], ("CY" if is_cy else "PY")

    # Case B: Match the single-year file that corresponds to the template sheet year
    if ys.year is not None:
        year_match = next((c for c in matches if c["year"] == ys.year), None)
        if year_match:
            return year_match["key"], "CY"

    # Default fallback to first match
    return matches[0]["key"], "CY"
