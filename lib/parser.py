"""QuickBooks export parser. Supports two variants:

  VARIANT A (single-year): one amount column on the right. Header on row 4 names
    the period (e.g., "Jan - Dec 25" or "Dec 31, 25"). Used by AP NE filings.

  VARIANT B (multi-year side-by-side): three amount columns on the right —
    Current Year, Prior Year, $ Change. Header on row 5 names them.
    Used by Popeyes IL filings ("Dec 31, 25" | "Dec 31, 24" | "$ Change").

Auto-detected per file. Each parsed row carries amount_cy, amount_py, change.
Downstream code reads amount_cy by default; legacy `amount` is an alias for CY.

Performance: files are parsed in parallel using a thread pool. File content is
hashed (SHA-256) so unchanged files are returned instantly from cache.
"""

import hashlib
import io
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Dict, Any

import openpyxl


# ── Module-level parse cache keyed by SHA-256 of file bytes ─────────────────
_PARSE_CACHE: Dict[str, Dict[str, Any]] = {}


def entity_from_filename(fn):
    base = os.path.basename(fn)
    m = re.match(r"^\d+\.\s*(.+?)(?:\s+Balance Sheet.*|\s+\d{4}-\d{4}|\.xlsx).*$", base, re.IGNORECASE)
    if m:
        return m.group(1).strip().rstrip(".")
    m = re.match(r"^\d+\.\s*(.+)\.xlsx$", base)
    return m.group(1).strip() if m else base.replace(".xlsx", "")


def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _detect_amount_columns(ws):
    period_pat = re.compile(r"(Jan\s*-\s*Dec|Dec\s*31|Jan\s*\d|Period|Change)", re.IGNORECASE)
    for r in range(3, 8):
        if r > ws.max_row:
            break
        hits = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and period_pat.search(v):
                hits.append((c, v.strip()))
        if hits:
            if len(hits) >= 3:
                cols = sorted([c for c, _ in hits])
                return ("triple", cols[0], cols[1], cols[2], r, r + 1)
            return ("single", hits[0][0], None, None, r, r + 1)
    return ("single", ws.max_column, None, None, 4, 5)


def parse_sheet(ws):
    variant, col_cy, col_py, col_change, header_row, data_start = _detect_amount_columns(ws)
    period_cy = ws.cell(row=header_row, column=col_cy).value if col_cy else None
    period_py = ws.cell(row=header_row, column=col_py).value if col_py else None
    period_cy = period_cy.strip() if isinstance(period_cy, str) else period_cy
    period_py = period_py.strip() if isinstance(period_py, str) else period_py

    rows = []
    hierarchy = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < data_start:
            continue

        amt_cy = amt_py = change = None
        if variant == "triple":
            if col_cy - 1 < len(row): amt_cy = row[col_cy - 1]
            if col_py - 1 < len(row): amt_py = row[col_py - 1]
            if col_change - 1 < len(row): change = row[col_change - 1]
            if not is_number(amt_cy): amt_cy = None
            if not is_number(amt_py): amt_py = None
            if not is_number(change): change = None
            label_search_end = min(col_cy - 1, col_py - 1, col_change - 1)
        else:
            label_search_end = len(row)
            for ci in range(len(row) - 1, -1, -1):
                if is_number(row[ci]):
                    amt_cy = row[ci]
                    label_search_end = ci
                    break

        lbl_idx, lbl = None, None
        for ci in range(label_search_end - 1, -1, -1):
            v = row[ci]
            if isinstance(v, str) and v.strip():
                lbl_idx, lbl = ci, v.strip()
                break

        if lbl is None:
            continue

        lbl_lower = lbl.lower()
        is_total = (
            lbl_lower.startswith("total ")
            or lbl_lower in ("net income", "net ordinary income", "net other income")
            or "total" in lbl_lower[:6]
        )
        is_section_only = amt_cy is None and amt_py is None

        while hierarchy and hierarchy[-1][0] >= lbl_idx:
            hierarchy.pop()

        parents = [p[1] for p in hierarchy]
        rows.append({
            "indent": lbl_idx,
            "label": lbl,
            "amount": amt_cy,
            "amount_cy": amt_cy,
            "amount_py": amt_py,
            "change": change,
            "is_total": is_total,
            "is_section_only": is_section_only,
            "breadcrumb": " > ".join(parents),
            "parents": parents,
            "row_idx": r_idx,
        })

        if not is_total:
            hierarchy.append((lbl_idx, lbl))

    return variant, period_cy, period_py, rows


def _parse_single_file(fname: str, file_bytes: bytes) -> tuple[str, dict]:
    """Parse one QB export file. Returns (entity_name, data_dict)."""
    entity = entity_from_filename(fname)

    # Check cache by SHA-256 hash
    digest = hashlib.sha256(file_bytes).hexdigest()
    if digest in _PARSE_CACHE:
        cached = dict(_PARSE_CACHE[digest])
        cached["file"] = fname  # filename may differ even for same content
        return entity, cached

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

        # Sheet classification — order matters: more specific first
        # Exclusion patterns for BS: "trial balance", "general ledger", "gl", "accrued"
        _EXCLUDE_FROM_BS = {"trial", "general ledger", "gl ", "accrued", "trial balance"}
        _EXCLUDE_FROM_PNL = {"trial", "general ledger", "gl ", "accrued"}

        pnl_candidates = []  # (priority, sheet_name)
        bs_candidates  = []

        for sn in wb.sheetnames:
            sn_l = sn.lower().strip()

            # Skip sheets that are clearly not financial statements
            if any(ex in sn_l for ex in ["trial balance", "general ledger", "accrued ledger", "gl "]):
                continue

            # P&L / Income Statement detection
            if any(k in sn_l for k in ["profit", "income statement", "income state"]):
                pnl_candidates.append((0, sn))  # exact keyword match = priority 0
            elif any(k in sn_l for k in ["p&l", "pnl", "operating"]):
                pnl_candidates.append((1, sn))
            elif "income" in sn_l and "balance" not in sn_l:
                pnl_candidates.append((2, sn))

            # Balance Sheet detection — must have "balance sheet" or just "balance" but NOT "trial"
            if "balance sheet" in sn_l:
                bs_candidates.append((0, sn))  # "Balance Sheet" = best match
            elif "balance" in sn_l and "trial" not in sn_l:
                bs_candidates.append((1, sn))

        pnl_sheet = min(pnl_candidates, key=lambda x: x[0])[1] if pnl_candidates else None
        bs_sheet  = min(bs_candidates,  key=lambda x: x[0])[1] if bs_candidates  else None

        if not pnl_sheet or not bs_sheet:
            result = {
                "file": fname,
                "error": "Missing P&L or BS sheet. Found: " + str(wb.sheetnames),
                "pnl_rows": [], "bs_rows": [],
                "pnl_period_cy": None, "pnl_period_py": None,
                "bs_period_cy": None, "bs_period_py": None,
                "variant": "unknown",
            }
            wb.close()
            return entity, result

        pnl_variant, pnl_cy, pnl_py, pnl_rows = parse_sheet(wb[pnl_sheet])
        bs_variant, bs_cy, bs_py, bs_rows = parse_sheet(wb[bs_sheet])
        variant = "triple" if "triple" in (pnl_variant, bs_variant) else "single"
        wb.close()

        result = {
            "file": fname,
            "variant": variant,
            "pnl_period_cy": pnl_cy, "pnl_period_py": pnl_py,
            "bs_period_cy": bs_cy, "bs_period_py": bs_py,
            "pnl_period": pnl_cy,
            "bs_period": bs_cy,
            "pnl_rows": pnl_rows,
            "bs_rows": bs_rows,
        }
        # Cache by content hash (without file-specific fields)
        cacheable = dict(result)
        cacheable.pop("file", None)
        _PARSE_CACHE[digest] = cacheable
        return entity, result

    except Exception as exc:
        return entity, {
            "file": fname,
            "error": str(exc),
            "pnl_rows": [], "bs_rows": [],
            "pnl_period_cy": None, "pnl_period_py": None,
            "bs_period_cy": None, "bs_period_py": None,
            "variant": "unknown",
        }


def parse_uploaded_files(uploaded_files, max_workers: int = 8):
    """Parse QB export files in parallel. Returns ordered dict matching upload order."""
    # Collect (fname, bytes) pairs
    tasks = []
    for uf in uploaded_files:
        if hasattr(uf, "name"):
            fname = uf.name
            file_bytes = uf.getvalue() if hasattr(uf, "getvalue") else uf.read()
        else:
            fname, file_bytes = uf
        tasks.append((fname, file_bytes))

    # Keep deterministic ordering that matches upload order
    results_ordered = [None] * len(tasks)

    with ThreadPoolExecutor(max_workers=min(max_workers, len(tasks))) as pool:
        future_to_idx = {
            pool.submit(_parse_single_file, fname, fbytes): idx
            for idx, (fname, fbytes) in enumerate(tasks)
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                entity, data = future.result()
            except Exception as exc:
                fname = tasks[idx][0]
                entity = entity_from_filename(fname)
                data = {
                    "file": fname, "error": str(exc),
                    "pnl_rows": [], "bs_rows": [],
                    "pnl_period_cy": None, "pnl_period_py": None,
                    "bs_period_cy": None, "bs_period_py": None,
                    "variant": "unknown",
                }
            results_ordered[idx] = (entity, data)

    all_data = {}
    for entity, data in results_ordered:
        if entity in all_data:
            # Deduplicate: append a suffix if same entity name appears twice
            i = 2
            while f"{entity} ({i})" in all_data:
                i += 1
            entity = f"{entity} ({i})"
        all_data[entity] = data

    return all_data
