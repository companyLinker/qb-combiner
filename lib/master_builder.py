"""Master consolidated workbook + CoA variants digest. Returns BytesIO objects."""

import io
import re
from collections import OrderedDict, defaultdict
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .mapping_rules import map_pnl, map_bs


HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOT_FILL = PatternFill("solid", fgColor="FFE699")
WARN_FILL = PatternFill("solid", fgColor="F4B084")
WHITE = Font(color="FFFFFF", bold=True)
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def copy_sheet_into_workbook(src_wb, sheet_name: str, dst_wb, new_name: str | None = None) -> None:
    """Copy a worksheet from src_wb into dst_wb.

    Copies cell values, number formats, basic styles (font/fill/border/alignment),
    merged-cell ranges, column widths, and row heights.
    The sheet is appended at the end of dst_wb.
    """
    src_ws = src_wb[sheet_name]

    # Determine a safe unique name (max 31 chars)
    desired = (new_name or sheet_name)[:31]
    name = desired
    i = 2
    while name in dst_wb.sheetnames:
        suffix = f" ({i})"
        name = desired[: 31 - len(suffix)] + suffix
        i += 1

    dst_ws = dst_wb.create_sheet(name)

    # Merged cells
    for merge in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge))

    # Column widths
    for col_letter, cd in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = cd.width

    # Row heights
    for row_idx, rd in src_ws.row_dimensions.items():
        if rd.height:
            dst_ws.row_dimensions[row_idx].height = rd.height

    # Cell values + styles
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            dst_cell.number_format = cell.number_format
            if cell.has_style:
                try:
                    dst_cell.font      = copy(cell.font)
                    dst_cell.fill      = copy(cell.fill)
                    dst_cell.border    = copy(cell.border)
                    dst_cell.alignment = copy(cell.alignment)
                except Exception:
                    pass



def style_header(ws, row=1):
    for c in ws[row]:
        c.font = WHITE; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX


def fit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_money(c):
    c.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def norm(s):
    if not s:
        return ""
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def build_pivot(rows_by_entity, entities):
    seen = OrderedDict()
    order_src = max(entities, key=lambda e: len(rows_by_entity[e]))

    def key(r): return (r["breadcrumb"], r["label"])

    for r in rows_by_entity[order_src]:
        seen[key(r)] = {"indent": r["indent"], "is_total": r["is_total"],
                        "is_section": r["is_section_only"],
                        "breadcrumb": r["breadcrumb"], "label": r["label"],
                        "amounts": {}, "amounts_cy": {}, "amounts_py": {}, "amounts_change": {}}
    for e in entities:
        for r in rows_by_entity[e]:
            k = key(r)
            if k not in seen:
                seen[k] = {"indent": r["indent"], "is_total": r["is_total"],
                           "is_section": r["is_section_only"],
                           "breadcrumb": r["breadcrumb"], "label": r["label"],
                           "amounts": {}, "amounts_cy": {}, "amounts_py": {}, "amounts_change": {}}
    for e in entities:
        for r in rows_by_entity[e]:
            k = key(r)
            cy = r.get("amount_cy") if r.get("amount_cy") is not None else r.get("amount")
            py = r.get("amount_py")
            change = r.get("change")
            
            if cy is not None:
                seen[k]["amounts_cy"][e] = seen[k]["amounts_cy"].get(e, 0) + (cy or 0)
                seen[k]["amounts"][e] = seen[k]["amounts_cy"][e]
            if py is not None:
                seen[k]["amounts_py"][e] = seen[k]["amounts_py"].get(e, 0) + (py or 0)
            if change is not None:
                seen[k]["amounts_change"][e] = seen[k]["amounts_change"].get(e, 0) + (change or 0)
    return seen


def write_pivot_sheets(wb, data, mapping_overrides=None, entity_mapping_overrides=None):
    entities = list(data.keys())
    overrides = mapping_overrides or {}
    entity_overrides = entity_mapping_overrides or {}

    pnl_pivot = build_pivot({e: data[e]["pnl_rows"] for e in entities}, entities)
    bs_pivot = build_pivot({e: data[e]["bs_rows"] for e in entities}, entities)

    def get_row_target_line(stmt_kind, bc, lbl):
        for e in entities:
            entity_key = f"E|{stmt_kind}|{e}|{bc}|{lbl}"
            if entity_key in entity_overrides and entity_overrides[entity_key]:
                return entity_overrides[entity_key]
        generic_key = f"{stmt_kind}|{bc}|{lbl}"
        if generic_key in overrides and overrides[generic_key]:
            return overrides[generic_key]
        if stmt_kind == "P&L":
            t, _ = map_pnl(bc, lbl)
            return t or ""
        else:
            t, _ = map_bs(bc, lbl)
            return t or ""

    def write_pivot(sheet_base_name, pivot, stmt_kind):
        tables_to_write = [
            ("CY", "amounts_cy", "Current Year"),
            ("PY", "amounts_py", "Prior Year"),
            ("Change", "amounts_change", "Change"),
        ]

        for suffix, key, title in tables_to_write:
            ws = wb.create_sheet(f"{sheet_base_name} {suffix}")
            
            # Write a title block for the table
            ws.append([f"{sheet_base_name} — {title}"])
            ws.cell(row=ws.max_row, column=1).font = Font(size=14, bold=True, color="1F3864")

            hdr = ["Breadcrumb", "Account / Label", "Target Line", "Indent", "Row Type"] + entities + ["Grand Total", "# Entities"]
            ws.append(hdr)
            hdr_row_idx = ws.max_row
            style_header(ws, row=hdr_row_idx)

            # Freeze panes for the table
            ws.freeze_panes = f"D{hdr_row_idx + 1}"

            for v in pivot.values():
                row_type = ("SectionOnly" if v["is_section"]
                            else ("Subtotal/Total" if v["is_total"] else "Leaf"))
                amts = v[key]
                values = [amts.get(e) for e in entities]
                total = sum(x for x in values if isinstance(x, (int, float)))
                n_rep = sum(1 for x in values if isinstance(x, (int, float)) and x != 0)
                
                tgt_line = "" if v["is_section"] or v["is_total"] else get_row_target_line(stmt_kind, v["breadcrumb"], v["label"])
                
                ws.append([v["breadcrumb"], v["label"], tgt_line, v["indent"], row_type] + values + [total, n_rep])
                rno = ws.max_row

                # Indent label
                ws.cell(row=rno, column=2).alignment = Alignment(indent=v["indent"])

                # Format money
                for ci in range(6, 6 + len(entities) + 1):
                    cell_val = ws.cell(row=rno, column=ci).value
                    if isinstance(cell_val, (int, float)):
                        fmt_money(ws.cell(row=rno, column=ci))

                if v["is_total"]:
                    for c in ws[rno]:
                        c.fill = TOT_FILL; c.font = Font(bold=True)
                elif v["is_section"]:
                    for c in ws[rno]:
                        c.fill = SUB_FILL; c.font = Font(bold=True, italic=True)

            fit(ws, [38, 38, 30, 7, 12] + [16] * len(entities) + [18, 12])

    write_pivot("P&L Pivot", pnl_pivot, "P&L")
    write_pivot("BS Pivot", bs_pivot, "BS")


def build_master_workbook(data, year_guess="2025", mapping_overrides=None, entity_mapping_overrides=None):
    """Build the consolidated master workbook. Returns BytesIO."""
    entities = list(data.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = f"MASTER CONSOLIDATED — QuickBooks P&L + Balance Sheet (FY {year_guess})"
    ws["A1"].font = Font(size=16, bold=True, color="1F3864")
    ws.merge_cells("A1:F1")
    
    readme = [
        ("", ""),
        ("Files consolidated", f"{len(entities)} entity Excel files"),
        ("Period (P&L)", data[entities[0]].get("pnl_period", "") if entities else ""),
        ("Period (BS)", data[entities[0]].get("bs_period", "") if entities else ""),
        ("", ""),
        ("Sheet", "What it contains"),
        ("P&L Long", "Every P&L line from every entity stacked. Long format for filter/pivot."),
        ("P&L Pivot CY", "P&L Accounts × entities matrix (Current Year) with Grand Total."),
        ("P&L Pivot PY", "P&L Accounts × entities matrix (Prior Year) with Grand Total."),
        ("P&L Pivot Change", "P&L Accounts × entities matrix ($ Change) with Grand Total."),
        ("BS Long", "Every BS line stacked."),
        ("BS Pivot CY", "BS Accounts × entities matrix (Current Year)."),
        ("BS Pivot PY", "BS Accounts × entities matrix (Prior Year)."),
        ("BS Pivot Change", "BS Accounts × entities matrix ($ Change)."),
        ("CoA Variants P&L", "Deduplicated P&L account list with frequency, breadcrumb, sum."),
        ("CoA Variants BS", "Same for Balance Sheet."),
    ]

    for i, (a, b) in enumerate(readme, 2):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120

    headers = ["Entity", "Source File", "Period", "Indent", "Row Type",
               "Section (L1)", "Category (L2)", "Sub-Category (L3)", "Sub-Sub (L4)", "Sub-Sub-Sub (L5)",
               "Account / Label", "Amount"]
    for stmt_label, stmt_key, period_key in [("P&L Long", "pnl_rows", "pnl_period"),
                                              ("BS Long", "bs_rows", "bs_period")]:
        ws = wb.create_sheet(stmt_label)
        ws.append(headers); style_header(ws); ws.freeze_panes = "A2"
        fit(ws, [28, 32, 16, 7, 12, 28, 28, 28, 28, 28, 38, 16])
        for entity in entities:
            info = data[entity]
            for r in info.get(stmt_key, []):
                p = r["parents"] + [""] * 5
                row_type = ("SectionOnly" if r["is_section_only"]
                            else ("Subtotal/Total" if r["is_total"] else "Leaf"))
                amt = r["amount"]
                ws.append([entity, info.get("file"), info.get(period_key), r["indent"], row_type,
                           p[0], p[1], p[2], p[3], p[4], r["label"], amt])
                cell = ws.cell(row=ws.max_row, column=12)
                if isinstance(amt, (int, float)): fmt_money(cell)
                if r["is_total"]:
                    for c in ws[ws.max_row]:
                        c.fill = TOT_FILL; c.font = Font(bold=True)
                elif r["is_section_only"]:
                    for c in ws[ws.max_row]:
                        c.fill = SUB_FILL; c.font = Font(bold=True, italic=True)

    # Write pivots using the helper
    write_pivot_sheets(wb, data, mapping_overrides, entity_mapping_overrides)

    pnl_pivot = build_pivot({e: data[e]["pnl_rows"] for e in entities}, entities)
    bs_pivot = build_pivot({e: data[e]["bs_rows"] for e in entities}, entities)

    # CoA Variants
    def write_coa(ws, pivot):
        hdr = ["#", "Section (L1)", "Category (L2)", "Sub (L3)", "Sub-Sub (L4)",
               "Account / Label", "Row Type", "Indent",
               "# Entities Using", "% of Entities", "Sum of Amount",
               "Entities Using", "Suggested Target (fill in)"]
        ws.append(hdr); style_header(ws); ws.freeze_panes = "A2"
        rows = []
        for v in pivot.values():
            parents = v["breadcrumb"].split(" > ") if v["breadcrumb"] else []
            p = parents + [""] * 4
            amounts = v["amounts"]
            used = [e for e in entities if e in amounts and amounts[e] != 0]
            row_type = ("SectionOnly" if v["is_section"]
                        else ("Subtotal/Total" if v["is_total"] else "Leaf"))
            rows.append([p[0], p[1], p[2], p[3], v["label"], row_type, v["indent"],
                         len(used), len(used) / len(entities), sum(amounts.get(e, 0) for e in entities),
                         ", ".join(used), ""])
        rows.sort(key=lambda r: (r[0] or "", r[1] or "", r[2] or "", r[3] or "", r[5] == "Subtotal/Total", r[4] or ""))
        for i, r in enumerate(rows, 1):
            ws.append([i] + r)
            rno = ws.max_row
            fmt_money(ws.cell(row=rno, column=11))
            ws.cell(row=rno, column=10).number_format = "0.0%"
        fit(ws, [5, 24, 24, 24, 24, 38, 14, 8, 12, 10, 20, 60, 36])

    write_coa(wb.create_sheet("CoA Variants P&L"), pnl_pivot)
    write_coa(wb.create_sheet("CoA Variants BS"), bs_pivot)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, pnl_pivot, bs_pivot


def build_variants_digest(data, pnl_pivot, bs_pivot):
    """Build the focused CoA variants digest workbook. Returns BytesIO."""
    entities = list(data.keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "CHART-OF-ACCOUNTS VARIANTS DIGEST"
    ws["A1"].font = Font(size=16, bold=True, color="1F3864")

    def leaf_index(pivot):
        idx = OrderedDict()
        for v in pivot.values():
            if v["is_section"] or v["is_total"]:
                continue
            idx[(v["breadcrumb"], v["label"])] = v
        return idx

    pnl_idx = leaf_index(pnl_pivot)
    bs_idx = leaf_index(bs_pivot)

    def variants_by_norm(idx):
        groups = defaultdict(list)
        for k, v in idx.items():
            groups[norm(v["label"])].append((k, v))
        return groups

    pnl_var = variants_by_norm(pnl_idx)
    bs_var = variants_by_norm(bs_idx)

    overview = [
        ("", ""),
        ("Files analyzed", f"{len(entities)} entities"),
        ("", ""),
        ("P&L unique LEAF accounts", len(pnl_idx)),
        ("BS unique LEAF accounts", len(bs_idx)),
        ("P&L spelling-variant groups", sum(1 for g in pnl_var.values() if len(g) > 1)),
        ("BS spelling-variant groups", sum(1 for g in bs_var.values() if len(g) > 1)),
    ]
    for i, (a, b) in enumerate(overview, 2):
        ws.cell(row=i, column=1, value=a).font = Font(bold=True)
        ws.cell(row=i, column=2, value=b)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 30

    def write_leaves(sheet_name, idx, var_groups):
        ws = wb.create_sheet(sheet_name)
        hdr = ["#", "Section (L1)", "Category (L2)", "Sub (L3)", "Sub-Sub (L4)",
               "Account Label", "Has Spelling Variants?",
               "# Entities Using", "% of Entities", "Sum All Entities",
               "Target Account (fill in)"]
        ws.append(hdr); style_header(ws); ws.freeze_panes = "A2"
        rows = []
        for v in idx.values():
            parents = v["breadcrumb"].split(" > ") if v["breadcrumb"] else []
            p = parents + [""] * 4
            amounts = v["amounts"]
            used = [e for e in entities if e in amounts and amounts[e] != 0]
            n_var = len(var_groups[norm(v["label"])])
            rows.append([p[0], p[1], p[2], p[3], v["label"], "YES" if n_var > 1 else "",
                         len(used), len(used) / len(entities), sum(amounts.get(e, 0) for e in entities)])
        rows.sort(key=lambda r: (r[0] or "", r[1] or "", r[2] or "", r[3] or "", r[4] or ""))
        for i, r in enumerate(rows, 1):
            ws.append([i] + r + [""])
            rno = ws.max_row
            fmt_money(ws.cell(row=rno, column=10))
            ws.cell(row=rno, column=9).number_format = "0.0%"
            if r[5] == "YES":
                ws.cell(row=rno, column=7).fill = WARN_FILL
                ws.cell(row=rno, column=7).font = Font(bold=True, color="9C0006")
        fit(ws, [5, 22, 22, 22, 22, 38, 14, 12, 10, 18, 36])

    write_leaves("P&L Leaf Accounts", pnl_idx, pnl_var)
    write_leaves("BS Leaf Accounts", bs_idx, bs_var)

    def write_spelling(sheet_name, idx, var_groups):
        ws = wb.create_sheet(sheet_name)
        hdr = ["Group #", "Normalized Account", "Variant Label", "Breadcrumb",
               "# Entities Using", "Entities Using", "Sum All Entities", "Likely Canonical"]
        ws.append(hdr); style_header(ws); ws.freeze_panes = "A2"
        grp_no = 0
        multi = sorted([(n, g) for n, g in var_groups.items() if len(g) > 1])
        for n, group in multi:
            grp_no += 1
            usage = [(k, v, sum(1 for e in entities if e in v["amounts"] and v["amounts"][e] != 0))
                     for k, v in group]
            usage.sort(key=lambda x: -x[2])
            canonical = usage[0][1]["label"]
            for k, v, n_used in usage:
                used = [e for e in entities if e in v["amounts"] and v["amounts"][e] != 0]
                total = sum(v["amounts"].get(e, 0) for e in entities)
                ws.append([grp_no, n, v["label"], v["breadcrumb"], n_used,
                           ", ".join(used), total, canonical])
                fmt_money(ws.cell(row=ws.max_row, column=7))
                for c in ws[ws.max_row]:
                    c.fill = WARN_FILL
        fit(ws, [8, 28, 30, 50, 12, 60, 16, 30])

    write_spelling("P&L Spelling Variants", pnl_idx, pnl_var)
    write_spelling("BS Spelling Variants", bs_idx, bs_var)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
